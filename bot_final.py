"""
Transfer Marketplace Bot — v4.1 (with Crypto Tracker)
================================
Features:
  - Arabic / English bilingual UI
  - Buy / Sell flow for USDT and USDC
  - Payment currencies: USD, EUR
  - Quick-amount buttons (50 – 1000) + custom entry
  - Offer rate entry per 100 units with live CoinGecko reference rate
  - Calculation preview with Confirm / Edit Rate before committing
  - Payment method: Bank Transfer / Cash / Mobile Wallet
  - Delivery method: Bank Transfer / Cash / Mobile Wallet / In-person
  - Optional email confirmation per trade
  - Full back-navigation at every step
  - Live exchange rates vs SYP (sp-today.com API)
  - Global forex rates (Frankfurter API)
  - Live gold prices all karats (sp-today.com API)
  - Admin panel  : /admin  — toggle currencies on/off
  - Admin stats  : /stats  — trade summary
  - Rate limiting (1.5 s between messages per user)
  - File + console logging  (bot.log)
  - Graceful shutdown on SIGTERM/SIGINT
  - 📊 Crypto Tracker: live prices, AI analysis, portfolio, Excel export
  - Posts confirmed trade to Telegram group with full user info
  - Sends confirmation email to user (Gmail SMTP)

Changes in v4.0:
  - [FIX] All secrets moved to .env file (no hardcoded credentials)
  - [FIX] All blocking HTTP/SMTP calls are now non-blocking (asyncio.to_thread)
  - [FIX] Buy flow now shows correct asset keyboard (asset_kb vs sell_asset_kb)
  - [FIX] Rate limiter uses time.monotonic() — no deprecated event loop call
  - [FIX] Email sending is now non-blocking (runs in thread)
  - [FIX] cb_back_to_asset uses correct keyboard for buy/sell

Dependencies:  pip install aiogram==2.25.1 requests beautifulsoup4 python-dotenv openpyxl openai
"""

import logging
import sqlite3
import time
import requests
import smtplib
import signal
import sys
import os
import asyncio
import re
from datetime import datetime
from email.mime.text import MIMEText
from collections import defaultdict
from dotenv import load_dotenv
from aiogram import Bot, Dispatcher, executor, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup
import io
from typing import Optional

# ══════════════════════════════════════════════
#  LOAD ENVIRONMENT VARIABLES
# ══════════════════════════════════════════════
load_dotenv()

# ══════════════════════════════════════════════
#  CONFIGURATION  —  loaded from .env file
# ══════════════════════════════════════════════
API_TOKEN  = os.getenv("API_TOKEN", "")
ADMIN_ID   = int(os.getenv("ADMIN_ID", "0"))
GROUP_ID   = int(os.getenv("GROUP_ID", "0"))
GMAIL_USER = os.getenv("GMAIL_USER", "")
GMAIL_PASS     = os.getenv("GMAIL_PASS", "")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
DATABASE_PATH  = os.getenv("DATABASE_PATH", "market.db")

if not API_TOKEN:
    raise RuntimeError("API_TOKEN is not set. Please add it to your .env file.")

# ══════════════════════════════════════════════
#  LOGGING  (file + console)
# ══════════════════════════════════════════════
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    handlers=[
        logging.FileHandler("bot.log", encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger("transfer_bot")

# ══════════════════════════════════════════════
#  CONSTANTS
# ══════════════════════════════════════════════
QUICK_AMOUNTS   = [50, 100, 150, 200, 250, 300, 500, 1000]
RATE_LIMIT_SECS = 1.5
EMAIL_RE        = re.compile(r"^[^\@\s]+@[^\@\s]+\.[^\@\s]+$")

PAYMENT_METHODS = {
    "ar": [
        ("🏦 تحويل بنكي",    "bank"),
        ("💵 نقداً",          "cash"),
        ("📱 محفظة موبايل",  "wallet"),
    ],
    "en": [
        ("🏦 Bank Transfer", "bank"),
        ("💵 Cash",          "cash"),
        ("📱 Mobile Wallet", "wallet"),
    ],
}

DELIVERY_METHODS = {
    "ar": [
        ("🏦 تحويل بنكي",    "bank"),
        ("💵 نقداً",          "cash"),
        ("📱 محفظة موبايل",  "wallet"),
        ("🤝 وجهاً لوجه",    "inperson"),
    ],
    "en": [
        ("🏦 Bank Transfer", "bank"),
        ("💵 Cash",          "cash"),
        ("📱 Mobile Wallet", "wallet"),
        ("🤝 In-person",     "inperson"),
    ],
}

# ══════════════════════════════════════════════
#  CRYPTO TRACKER — CONSTANTS & TOKEN METADATA
# ══════════════════════════════════════════════
COINGECKO_BASE = "https://api.coingecko.com/api/v3"

TOKEN_EMOJIS = {
    "bitcoin": "🟠", "ethereum": "🔵", "solana": "🟣",
    "tether": "💚", "binancecoin": "🔴", "ripple": "🔵",
    "cardano": "🔵", "dogecoin": "🟡", "polkadot": "🟣",
    "chainlink": "🔵", "litecoin": "⚫", "avalanche-2": "🔴",
    "matic-network": "🟣", "uniswap": "🩷", "cosmos": "🔵",
    "tron": "🔴", "shiba-inu": "🟡", "near": "🟢",
    "stellar": "🔵", "monero": "🟠",
}

POPULAR_TOKENS = [
    ("bitcoin", "BTC"),      ("ethereum", "ETH"),    ("solana", "SOL"),
    ("tether", "USDT"),      ("binancecoin", "BNB"), ("ripple", "XRP"),
    ("cardano", "ADA"),      ("dogecoin", "DOGE"),   ("polkadot", "DOT"),
    ("chainlink", "LINK"),   ("litecoin", "LTC"),    ("avalanche-2", "AVAX"),
    ("matic-network", "MATIC"), ("uniswap", "UNI"), ("cosmos", "ATOM"),
    ("tron", "TRX"),         ("shiba-inu", "SHIB"),  ("near", "NEAR"),
    ("stellar", "XLM"),      ("monero", "XMR"),
]

# ══════════════════════════════════════════════
#  RATE LIMITER
#  FIX: uses time.monotonic() instead of the deprecated
#       asyncio.get_event_loop().time()
# ══════════════════════════════════════════════
_last_seen: dict = defaultdict(float)

def is_rate_limited(user_id: int) -> bool:
    now = time.monotonic()
    if now - _last_seen[user_id] < RATE_LIMIT_SECS:
        return True
    _last_seen[user_id] = now
    return False

# ══════════════════════════════════════════════
#  DATABASE
# ══════════════════════════════════════════════
def init_db() -> None:
    with sqlite3.connect(DATABASE_PATH) as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS trades (
                id        INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id   TEXT,    username TEXT,   email    TEXT,
                type      TEXT,    asset    TEXT,
                amount    REAL,    price    REAL,
                currency  TEXT,    payment  TEXT,
                delivery  TEXT,    total    REAL,
                usd_snap  REAL,    eur_snap REAL,
                status    TEXT,    date     TEXT
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                item    TEXT PRIMARY KEY,
                enabled INTEGER DEFAULT 1
            )
        """)
        conn.execute("INSERT OR IGNORE INTO settings VALUES ('USD', 1)")
        conn.execute("INSERT OR IGNORE INTO settings VALUES ('EUR', 1)")
        conn.execute("INSERT OR IGNORE INTO settings VALUES ('SYP', 1)")
        # ── Section toggles ──────────────────────────────────
        conn.execute("INSERT OR IGNORE INTO settings VALUES ('section_buy',    1)")
        conn.execute("INSERT OR IGNORE INTO settings VALUES ('section_sell',   1)")
        conn.execute("INSERT OR IGNORE INTO settings VALUES ('section_rates',  1)")
        conn.execute("INSERT OR IGNORE INTO settings VALUES ('section_gold',   1)")
        conn.execute("INSERT OR IGNORE INTO settings VALUES ('section_crypto', 1)")
        conn.commit()
    # ── Crypto Tracker tables ──────────────────────────────
    with sqlite3.connect(DATABASE_PATH) as tconn:
        tconn.execute("""
            CREATE TABLE IF NOT EXISTS tracked_tokens (
                id       TEXT PRIMARY KEY,
                symbol   TEXT NOT NULL,
                enabled  INTEGER DEFAULT 1,
                added_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)
        tconn.execute("""
            CREATE TABLE IF NOT EXISTS user_portfolio (
                user_id    INTEGER NOT NULL,
                token_id   TEXT NOT NULL,
                quantity   REAL DEFAULT 0,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (user_id, token_id)
            )
        """)
        tconn.execute("""
            CREATE TABLE IF NOT EXISTS user_lang (
                user_id  INTEGER PRIMARY KEY,
                lang     TEXT DEFAULT 'en'
            )
        """)
        tconn.commit()

def save_user_lang(user_id: int, lang: str) -> None:
    """Persist user language preference (ar/en) to DB."""
    with sqlite3.connect(DATABASE_PATH) as conn:
        conn.execute(
            "INSERT OR REPLACE INTO user_lang (user_id, lang) VALUES (?, ?)",
            (user_id, lang)
        )
        conn.commit()

def get_user_lang(user_id: int) -> str:
    """Return stored language preference for user, default 'en'."""
    with sqlite3.connect(DATABASE_PATH) as conn:
        row = conn.execute(
            "SELECT lang FROM user_lang WHERE user_id = ?", (user_id,)
        ).fetchone()
    return row[0] if row else "en"

def get_enabled_currencies() -> list:
    with sqlite3.connect(DATABASE_PATH) as conn:
        rows = conn.execute(
            "SELECT item FROM settings WHERE enabled=1 ORDER BY item"
        ).fetchall()
    return [r[0] for r in rows]

def get_all_settings() -> list:
    with sqlite3.connect(DATABASE_PATH) as conn:
        return conn.execute("SELECT item, enabled FROM settings ORDER BY item").fetchall()

def toggle_setting(item: str) -> int:
    with sqlite3.connect(DATABASE_PATH) as conn:
        cur = conn.execute("SELECT enabled FROM settings WHERE item=?", (item,)).fetchone()
        new_val = 0 if cur[0] else 1
        conn.execute("UPDATE settings SET enabled=? WHERE item=?", (new_val, item))
        conn.commit()
    return new_val

def is_section_enabled(section: str) -> bool:
    """Return True if the given section is open, False if admin closed it."""
    with sqlite3.connect(DATABASE_PATH) as conn:
        row = conn.execute(
            "SELECT enabled FROM settings WHERE item=?", (section,)
        ).fetchone()
    return bool(row[0]) if row else True

SECTION_CLOSED_MSG = {
    "section_buy":    "🔴 قسم الشراء مغلق مؤقتًا.\n🔴 Buy section is currently closed.",
    "section_sell":   "🔴 قسم البيع مغلق مؤقتًا.\n🔴 Sell section is currently closed.",
    "section_rates":  "🔴 قسم أسعار الصرف مغلق مؤقتًا.\n🔴 Exchange Rates section is currently closed.",
    "section_gold":   "🔴 قسم أسعار الذهب مغلق مؤقتًا.\n🔴 Gold Prices section is currently closed.",
    "section_crypto": "🔴 قسم العملات الرقمية مغلق مؤقتًا.\n🔴 Crypto Tracker section is currently closed.",
}

# ──────────────────────────────────────────────
#  FIX: get_market_rates is now a plain sync function
#       called only inside save_trade (which is itself
#       wrapped in asyncio.to_thread in the handler).
# ──────────────────────────────────────────────
def get_market_rates() -> tuple:
    """Fetch live USD→EUR rate. Returns (1.0, eur_rate) or (0.0, 0.0) on failure."""
    try:
        data = requests.get(
            "https://api.frankfurter.app/latest?from=USD&to=EUR", timeout=5
        ).json()
        return 1.0, data["rates"]["EUR"]
    except Exception:
        try:
            data = requests.get(
                "https://api.exchangerate.host/latest?base=USD&symbols=EUR", timeout=5
            ).json()
            return 1.0, data["rates"]["EUR"]
        except Exception:
            return 0.0, 0.0

# ══════════════════════════════════════════════
#  CURRENCY FLAGS MAP
# ══════════════════════════════════════════════
CURRENCY_FLAGS = {
    "USD": "🇺🇸", "EUR": "🇪🇺", "TRY": "🇹🇷", "SAR": "🇸🇦",
    "AED": "🇦🇪", "EGP": "🇪🇬", "GBP": "🇬🇧", "QAR": "🇶🇦",
    "KWD": "🇰🇼", "JOD": "🇯🇴", "BHD": "🇧🇭", "OMR": "🇴🇲",
    "CAD": "🇨🇦", "AUD": "🇦🇺", "CHF": "🇨🇭", "SEK": "🇸🇪",
    "NOK": "🇳🇴", "DKK": "🇩🇰", "RUB": "🇷🇺", "LYD": "🇱🇾",
    "NZD": "🇳🇿", "SGD": "🇸🇬", "MYR": "🇲🇾", "BRL": "🇧🇷",
    "ZAR": "🇿🇦", "MAD": "🇲🇦", "DZD": "🇩🇿", "TND": "🇹🇳",
    "IQD": "🇮🇶", "IRR": "🇮🇷",
}

# ══════════════════════════════════════════════
#  SP-TODAY SCRAPER
#  FIX: called via asyncio.to_thread() in handlers
# ══════════════════════════════════════════════
def get_sptoday_rates() -> list:
    """
    Fetch live exchange rates from sp-today API.
    Returns list of dicts with keys: code, name_ar, flag, buy, sell,
    change_day, change_week, change_month, change_year, day_high, day_low.
    """
    try:
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            "Accept": "application/json",
            "Origin": "https://sp-today.com",
            "Referer": "https://sp-today.com/currencies",
        }
        resp = requests.get(
            "https://api-v2.sp-today.com/api/v1/overview?lang=ar&city=damascus",
            headers=headers, timeout=10
        )
        payload = resp.json()
        raw_rates = payload["data"]["rates"]

        result = []
        for item in raw_rates:
            city = item.get("cities", {}).get("damascus", {})
            if not city:
                continue
            result.append({
                "code":         item.get("code", "???"),
                "name_ar":      item.get("name_ar", item.get("name", "")),
                "flag":         item.get("flag", "💵"),
                "buy":          city.get("buy", 0),
                "sell":         city.get("sell", 0),
                "change_day":   city.get("change", 0),
                "change_week":  city.get("change_week", 0),
                "change_month": city.get("change_month", 0),
                "change_year":  city.get("change_year", 0),
                "day_high":     city.get("day_high", 0),
                "day_low":      city.get("day_low", 0),
            })

        # Fetch extra currencies not in the overview (e.g. SEK)
        extra_currencies = ["SEK"]
        for code in extra_currencies:
            try:
                r2 = requests.get(
                    f"https://api-v2.sp-today.com/api/v1/currency/{code}?city=damascus",
                    headers=headers, timeout=10
                )
                d = r2.json()
                item = d["data"]["currency"]
                city = item.get("cities", {}).get("damascus", {})
                if city:
                    result.append({
                        "code":         item.get("code", code),
                        "name_ar":      item.get("name_ar", item.get("name", code)),
                        "flag":         item.get("flag", "🏳️"),
                        "buy":          city.get("buy", 0),
                        "sell":         city.get("sell", 0),
                        "change_day":   city.get("change", 0),
                        "change_week":  city.get("change_week", 0),
                        "change_month": city.get("change_month", 0),
                        "change_year":  city.get("change_year", 0),
                        "day_high":     city.get("day_high", 0),
                        "day_low":      city.get("day_low", 0),
                    })
            except Exception as e:
                logger.warning("Could not fetch extra currency %s: %s", code, e)

        logger.info("sp-today API: fetched %d currency rates", len(result))
        return result

    except Exception as exc:
        logger.error("get_sptoday_rates: %s", exc)
        return []


def get_global_rates() -> list:
    """
    Fetch global forex rates for SEK/EUR, SEK/USD, USD/EUR from Frankfurter API.
    Returns list of dicts with keys: base, quote, flag_base, flag_quote,
    mid, buy, sell, change_day, change_week, day_high, day_low.
    """
    try:
        from datetime import timedelta
        pairs = [
            ("SEK", "EUR", "🇸🇪", "🇪🇺", 0.003),
            ("SEK", "USD", "🇸🇪", "🇺🇸", 0.003),
            ("USD", "EUR", "🇺🇸", "🇪🇺", 0.0015),
        ]
        today     = datetime.now()
        yesterday = (today - timedelta(days=1)).strftime("%Y-%m-%d")
        week_ago  = (today - timedelta(days=7)).strftime("%Y-%m-%d")

        def _fetch(date_str, base, quote):
            url = f"https://api.frankfurter.app/{date_str}?from={base}&to={quote}"
            r = requests.get(url, timeout=10)
            return r.json()["rates"][quote]

        result = []
        for base, quote, flag_b, flag_q, spread in pairs:
            try:
                mid_today = _fetch("latest", base, quote)
                try:
                    mid_prev  = _fetch(yesterday, base, quote)
                    chg_day   = ((mid_today - mid_prev) / mid_prev) * 100
                except Exception:
                    chg_day = 0.0
                try:
                    mid_week  = _fetch(week_ago, base, quote)
                    chg_week  = ((mid_today - mid_week) / mid_week) * 100
                except Exception:
                    chg_week = 0.0

                buy  = mid_today * (1 - spread / 2)
                sell = mid_today * (1 + spread / 2)
                result.append({
                    "base":       base,
                    "quote":      quote,
                    "flag_base":  flag_b,
                    "flag_quote": flag_q,
                    "mid":        mid_today,
                    "buy":        buy,
                    "sell":       sell,
                    "change_day": round(chg_day, 4),
                    "change_week": round(chg_week, 4),
                    "day_high":   mid_today * (1 + spread / 2),
                    "day_low":    mid_today * (1 - spread / 2),
                })
            except Exception as e:
                logger.warning("Global rate %s/%s error: %s", base, quote, e)

        logger.info("Frankfurter: fetched %d global pairs", len(result))
        return result

    except Exception as exc:
        logger.error("get_global_rates: %s", exc)
        return []


def format_global_rates_message(pairs: list, lang: str) -> str:
    """Format global forex rates into the same card design as SYP rates."""
    if not pairs:
        return (
            "⚠️ تعذّر تحميل الأسعار العالمية حالياً. حاول مرة أخرى لاحقاً."
            if lang == "ar" else
            "⚠️ Could not load global rates right now. Please try again later."
        )

    now = datetime.now().strftime("%H:%M")
    if lang == "ar":
        title  = f"🌍 <b>أسعار الصرف العالمية</b>  |  لكل 1000 وحدة\n🕐 آخر تحديث: {now}"
        buy_l  = "شراء"; sell_l = "بيع"; high_l = "أعلى"; low_l = "أدنى"
    else:
        title  = f"🌍 <b>Global Exchange Rates</b>  |  per 1000 units\n🕐 Last updated: {now}"
        buy_l  = "Buy"; sell_l = "Sell"; high_l = "High"; low_l = "Low"

    # Name labels
    names_ar = {"SEK/EUR": "كرون ← يورو", "SEK/USD": "كرون ← دولار", "USD/EUR": "دولار ← يورو"}
    names_en = {"SEK/EUR": "SEK → EUR", "SEK/USD": "SEK → USD", "USD/EUR": "USD → EUR"}

    lines = [title, ""]
    for p in pairs:
        key       = f"{p['base']}/{p['quote']}"
        name      = names_ar.get(key, key) if lang == "ar" else names_en.get(key, key)
        flag      = f"{p['flag_base']}{p['flag_quote']}"
        buy       = f"{p['buy'] * 1000:.4f}"
        sell      = f"{p['sell'] * 1000:.4f}"
        high      = f"{p['day_high'] * 1000:.4f}"
        low       = f"{p['day_low'] * 1000:.4f}"
        chg_day   = p["change_day"]
        chg_week  = p["change_week"]

        if chg_day > 0:
            day_badge = f"📈 +{chg_day:.2f}%"
        elif chg_day < 0:
            day_badge = f"📉 {chg_day:.2f}%"
        else:
            day_badge = "➖ —"

        lines.append(f"┌ {p['flag_base']} <code>1000 {p['base']}</code> ➡️ {p['flag_quote']} <b>{p['quote']}</b>  {day_badge}")
        lines.append(f"│  💰 <b>{buy_l}:</b>  <code>{buy}</code>   <b>{sell_l}:</b>  <code>{sell}</code>")
        lines.append(f"│  📊 <b>{high_l}:</b> <code>{high}</code>   <b>{low_l}:</b> <code>{low}</code>")
        lines.append(f"└  {_fmt_change(chg_week, 'أسبوعي', 'Week', lang)}")
        lines.append("")

    source = (
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "📌 <a href='https://www.frankfurter.app'>Frankfurter.app</a>"
    )
    lines.append(source)
    return "\n".join(lines)


def get_gold_rates() -> list:
    """
    Fetch live gold prices from sp-today API.
    Returns list of dicts with keys: karat, buy, sell,
    change_day, change_week, change_month, change_year, day_high, day_low.
    """
    try:
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            "Accept": "application/json",
            "Origin": "https://sp-today.com",
            "Referer": "https://sp-today.com/gold",
        }
        resp = requests.get(
            "https://api-v2.sp-today.com/api/v1/gold",
            headers=headers, timeout=10
        )
        payload = resp.json()
        raw_karats = payload["data"]["karats"]

        result = []
        for item in raw_karats:
            city = item.get("cities", {}).get("damascus", {})
            if not city:
                continue
            result.append({
                "karat":        item.get("karat", "??K"),
                "buy":          city.get("buy", 0),
                "sell":         city.get("sell", 0),
                "change_day":   city.get("change", 0),
                "change_week":  city.get("change_week", 0),
                "change_month": city.get("change_month", 0),
                "change_year":  city.get("change_year", 0),
                "day_high":     city.get("day_high", 0),
                "day_low":      city.get("day_low", 0),
            })

        logger.info("sp-today gold API: fetched %d karats", len(result))
        return result

    except Exception as exc:
        logger.error("get_gold_rates: %s", exc)
        return []


def format_gold_message(gold: list, lang: str) -> str:
    """Format gold rates into a clean Telegram HTML message."""
    if not gold:
        return (
            "⚠️ تعذّر تحميل أسعار الذهب حالياً. حاول مرة أخرى لاحقاً."
            if lang == "ar" else
            "⚠️ Could not load gold prices right now. Please try again later."
        )

    now = datetime.now().strftime("%H:%M")
    if lang == "ar":
        title = f"🥇 <b>أسعار الذهب في سوريا (ل.س)</b>\n🕐 آخر تحديث: {now}"
        buy_l  = "شراء"
        sell_l = "بيع"
        high_l = "أعلى"
        low_l  = "أدنى"
        gram_l = "غرام"
    else:
        title = f"🥇 <b>Gold Prices in Syria (SYP)</b>\n🕐 Last updated: {now}"
        buy_l  = "Buy"
        sell_l = "Sell"
        high_l = "High"
        low_l  = "Low"
        gram_l = "gram"

    lines = [title, ""]

    for g in gold:
        karat  = g["karat"]
        buy    = f"{g['buy']:,}"
        sell   = f"{g['sell']:,}"
        high   = f"{g['day_high']:,}"
        low    = f"{g['day_low']:,}"
        chg_day   = g["change_day"]
        chg_week  = g["change_week"]

        # Daily badge
        if chg_day > 0:
            day_badge = f"📈 +{chg_day:.2f}%"
        elif chg_day < 0:
            day_badge = f"📉 {chg_day:.2f}%"
        else:
            day_badge = "➖ —"

        lines.append(f"┌ 🥇 <b>{karat}</b>  ({gram_l})  {day_badge}")
        lines.append(f"│  💰 <b>{buy_l}:</b>  <code>{buy}</code>   <b>{sell_l}:</b>  <code>{sell}</code>")
        lines.append(f"│  📊 <b>{high_l}:</b> <code>{high}</code>   <b>{low_l}:</b> <code>{low}</code>")
        lines.append(f"└  {_fmt_change(chg_week, 'أسبوعي', 'Week', lang)}")
        lines.append("")  # blank line between karats

    source = (
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "📌 <a href='https://sp-today.com/gold'>sp-today.com/gold</a>"
    )
    lines.append(source)
    return "\n".join(lines)


def _fmt_change(val: float, label_ar: str, label_en: str, lang: str) -> str:
    """Format a change percentage value with arrow, color indicator and label."""
    label = label_ar if lang == "ar" else label_en
    if val > 0:
        return f"📈 <b>{label}:</b> <i>+{val:.2f}%</i>"
    elif val < 0:
        return f"📉 <b>{label}:</b> <i>{val:.2f}%</i>"
    else:
        return f"➖ <b>{label}:</b> <i>0.00%</i>"


def format_rates_message(rates: list, lang: str) -> str:
    """Format the API rates into a clean Telegram HTML message."""
    if not rates:
        return (
            "⚠️ تعذّر تحميل الأسعار حالياً. حاول مرة أخرى لاحقاً."
            if lang == "ar"
            else "⚠️ Could not load rates right now. Please try again later."
        )

    now = datetime.now().strftime("%H:%M")
    if lang == "ar":
        title    = f"💹 <b>أسعار الصرف مقابل الليرة السورية</b>\n🕐 آخر تحديث: {now}"
        buy_l    = "شراء"
        sell_l   = "بيع"
        high_l   = "أعلى"
        low_l    = "أدنى"
    else:
        title    = f"💹 <b>Exchange Rates vs Syrian Pound (SYP)</b>\n🕐 Last updated: {now}"
        buy_l    = "Buy"
        sell_l   = "Sell"
        high_l   = "High"
        low_l    = "Low"

    # Custom base denomination per currency (default 100, SEK uses 1000)
    BASE_DENOM = {"SEK": 1000}

    # Desired order: USD, EUR, SEK first, then the rest
    ORDER = ["USD", "EUR", "SEK"]
    ordered = sorted(rates, key=lambda r: (
        ORDER.index(r["code"]) if r["code"] in ORDER else len(ORDER) + 1
    ))

    lines = [title, ""]

    for r in ordered:
        flag         = r["flag"]
        code         = r["code"]
        name_display = r["name_ar"] if lang == "ar" else code
        denom        = BASE_DENOM.get(code, 100)
        multiplier   = denom / 100  # scale buy/sell/high/low
        buy          = f"{int(r['buy'] * multiplier):,}"
        sell         = f"{int(r['sell'] * multiplier):,}"
        high         = f"{int(r['day_high'] * multiplier):,}"
        low          = f"{int(r['day_low'] * multiplier):,}"
        chg_day      = r["change_day"]
        chg_week     = r["change_week"]
        chg_month    = r["change_month"]
        chg_year     = r.get("change_year", 0)

        # Daily badge: show change if non-zero, else show dash
        if chg_day > 0:
            day_badge = f"📈 +{chg_day:.2f}%"
        elif chg_day < 0:
            day_badge = f"📉 {chg_day:.2f}%"
        else:
            day_badge = "➖ —"
        lines.append(f"┌ {flag} <b>{code}</b>  {name_display}  <code>{denom}</code>  {day_badge}")
        lines.append(f"│  💰 <b>{buy_l}:</b>  <code>{buy}</code>   <b>{sell_l}:</b>  <code>{sell}</code>")
        lines.append(f"│  📊 <b>{high_l}:</b> <code>{high}</code>   <b>{low_l}:</b> <code>{low}</code>")
        lines.append(f"└  {_fmt_change(chg_week, 'أسبوعي', 'Week', lang)}")
        lines.append("")   # blank line between currencies

    source = (
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "📌 <a href='https://sp-today.com/currencies'>sp-today.com</a>"
    )
    lines.append(source)
    return "\n".join(lines)


def save_trade(data: dict, user: types.User) -> tuple:
    total = (data["amount"] / 100) * data["price"]
    date  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    usd_snap, eur_snap = get_market_rates()
    with sqlite3.connect(DATABASE_PATH) as conn:
        cur = conn.execute(
            """INSERT INTO trades
               (user_id, username, email, type, asset, amount, price, currency,
                payment, delivery, total, usd_snap, eur_snap, status, date)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                str(user.id), user.username or "", data.get("email", ""),
                data["type"], data["asset"], data["amount"], data["price"],
                data["currency"], data.get("payment", ""), data.get("delivery", ""),
                total, usd_snap, eur_snap, "PENDING", date,
            ),
        )
        trade_id = cur.lastrowid
        conn.commit()
        row = conn.execute("SELECT * FROM trades WHERE id=?", (trade_id,)).fetchone()
    return row

def get_stats() -> dict:
    with sqlite3.connect(DATABASE_PATH) as conn:
        total     = conn.execute("SELECT COUNT(*) FROM trades").fetchone()[0]
        buy_cnt   = conn.execute("SELECT COUNT(*) FROM trades WHERE type='buy'").fetchone()[0]
        sell_cnt  = conn.execute("SELECT COUNT(*) FROM trades WHERE type='sell'").fetchone()[0]
        pending   = conn.execute("SELECT COUNT(*) FROM trades WHERE status='PENDING'").fetchone()[0]
        completed = conn.execute("SELECT COUNT(*) FROM trades WHERE status='COMPLETED'").fetchone()[0]
        volumes   = conn.execute(
            "SELECT asset, SUM(amount) FROM trades GROUP BY asset"
        ).fetchall()
    return {
        "total": total, "buy": buy_cnt, "sell": sell_cnt,
        "pending": pending, "completed": completed, "volumes": volumes,
    }

# ══════════════════════════════════════════════
#  EMAIL
#  FIX: send_confirmation_email is a plain sync function.
#       It is called via asyncio.to_thread() in _finish_trade
#       so it never blocks the event loop.
# ══════════════════════════════════════════════
def send_confirmation_email(recipient: str, row: tuple) -> None:
    # row cols: id, user_id, username, email, type, asset, amount, price,
    #           currency, payment, delivery, total, usd_snap, eur_snap, status, date
    try:
        rate_line = f"1 USD = {row[13]:.4f} EUR" if row[13] else "N/A"
        body = (
            f"تأكيد الصفقة / Trade Confirmation #{row[0]}\n\n"
            f"النوع / Type      : {row[4]}\n"
            f"العملة / Asset    : {row[5]}\n"
            f"المبلغ / Amount   : {row[6]:,.2f}\n"
            f"السعر / Price     : {row[7]:,.2f} {row[8]}\n"
            f"الدفع / Payment   : {row[9]}\n"
            f"التسليم / Deliver : {row[10]}\n"
            f"الإجمالي / Total  : {row[11]:,.2f} {row[8]}\n"
            f"سعر السوق / Rate  : {rate_line}\n"
            f"الحالة / Status   : {row[14]}\n"
            f"التاريخ / Date    : {row[15]}\n"
        )
        msg = MIMEText(body, "plain", "utf-8")
        msg["Subject"] = f"إيصال صفقة / Trade Receipt #{row[0]}"
        msg["From"]    = GMAIL_USER
        msg["To"]      = recipient
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(GMAIL_USER, GMAIL_PASS)
            server.send_message(msg)
        logger.info("Confirmation email sent to %s", recipient)
    except Exception as exc:
        logger.error("Email error: %s", exc)

# ══════════════════════════════════════════════
#  BOT + DISPATCHER
# ══════════════════════════════════════════════
bot = Bot(token=API_TOKEN, parse_mode="HTML")
dp  = Dispatcher(bot, storage=MemoryStorage())

# ══════════════════════════════════════════════
#  FSM STATES
# ══════════════════════════════════════════════
class MarketFlow(StatesGroup):
    language     = State()
    type         = State()
    asset        = State()
    amount       = State()
    currency     = State()
    price        = State()
    confirm_calc = State()   # show calculation preview before payment
    payment      = State()   # payment method (bank/cash/wallet)
    delivery     = State()   # delivery method (bank/cash/wallet/inperson)
    email        = State()

# ══════════════════════════════════════════════
#  LOCALISED STRINGS
# ══════════════════════════════════════════════
TX = {
    "ar": {
        "welcome": (
            "🌟 <b>مرحبًا بك في بوت التحويل!</b> 🌟\n\n"
            "نسعى لجعل تجربة التحويل لديك سلسة، آمنة وذكية.\n\n"
            "💰 اختر العملة التي ترغب في استلامها\n"
            "💳 واختر طريقة الدفع المفضلة لديك\n\n"
            "للبدء، اضغط على /start واستعد لتجربة تحويل فريدة وسريعة! 🚀\n\n"
            "اختر لغتك للبدء 👇\n"
            "<i>Select your language to begin 👇</i>"
        ),
        "select_type"     : "👋 <b>أهلاً! كيف يمكنني مساعدتك اليوم؟</b>\n\nاختر ما تريد القيام به:",
        "btn_buy"         : "🛒 شراء",
        "btn_sell"        : "💸 بيع",
        "select_asset_b"  : "💱 <b>رائع! أي عملة تود شراءها؟</b>\n\nاختر من الخيارات أدناه 👇",
        "select_asset_sl" : "💱 <b>حسناً! أي عملة تود بيعها؟</b>\n\nاختر من الخيارات أدناه 👇",
        "enter_amount"    : "💵 <b>كم تريد؟</b>\n\nاختر مبلغاً سريعاً أدناه 👇 أو اضغط <b>أخرى</b> لإدخال مبلغ مختلف.",
        "custom_amount"   : "✏️ <b>أدخل المبلغ:</b>",
        "other_amount"    : "✍️ <b>أدخل المبلغ الذي تريده:</b>\n\nاكتب الرقم وأرسله 👇",
        "select_pay_cur"  : "🌍 <b>ما هي عملة الدفع؟</b>\n\nاختر العملة التي ستدفع بها 👇",
        "enter_price"     : "💱 <b>ما هو سعرك المطلوب؟</b>\n\nاكتب المبلغ وأرسله 👇",
        "select_payment"  : "💳 <b>كيف تريد الدفع؟</b>\n\nاختر طريقة الدفع المناسبة لك 👇",
        "select_delivery" : "🚚 <b>كيف تريد استلام مبلغك؟</b>\n\nاختر طريقة الاستلام 👇",
        "enter_email"     : "📧 <b>هل تريد إرسال تأكيد على بريدك الإلكتروني؟</b>\n\nأدخل بريدك الإلكتروني أو اضغط <b>تخطي</b> إذا لم تكن بحاجة لذلك.",
        "btn_skip_email"  : "⏭️ تخطي",
        "btn_rates"       : "💹 أسعار الصرف",
        "btn_gold"        : "🥇 أسعار الذهب",
        "btn_back"        : "🏠 العودة للقائمة",
        "bad_amount"      : "⚠️ <b>المبلغ غير صحيح.</b>\n\nالرجاء إدخال رقم موجب وحاول مرة أخرى 👇",
        "bad_email"       : "⚠️ <b>البريد الإلكتروني غير صحيح.</b>\n\nحاول مرة أخرى أو اضغط <b>تخطي</b>.",
        "about": (
            "ℹ️ <b>عن البوت</b>\n\n"
            "بوت التحويل v4.0\n"
            "منصة آمنة وسريعة لتحويل العملات.\n\n"
            "الأوامر المتاحة:\n"
            "/start — بدء تحويل جديد\n"
            "/about — معلومات عن البوت\n\n"
            "للدعم تواصل مع المشرف."
        ),
        "type_lbl"     : "النوع",
        "asset_lbl"    : "العملة",
        "amount_lbl"   : "المبلغ",
        "price_lbl"    : "السعر",
        "payment_lbl"  : "طريقة الدفع",
        "delivery_lbl" : "طريقة التسليم",
        "total_lbl"    : "الإجمالي",
        "rate_lbl"     : "سعر السوق",
        "date_lbl"     : "التاريخ",
        "done_title"   : "🎉 <b>تم تسجيل طلبك بنجاح!</b>\n\nشكراً لك، سيتواصل معك فريقنا قريباً. إليك ملخص طلبك:",
    },
    "en": {
        "welcome": (
            "🌟 <b>Welcome to the Transfer Bot!</b> 🌟\n\n"
            "We make your transfer experience smooth, secure and smart.\n\n"
            "💰 Choose the currency you want to receive\n"
            "💳 Choose your preferred payment method\n\n"
            "Press /start to begin a fast and unique transfer experience! 🚀\n\n"
            "<i>اختر لغتك للبدء 👇</i>\n"
            "Select your language to begin 👇"
        ),
        "select_type"     : "👋 <b>Hey there! What would you like to do today?</b>\n\nChoose an option below:",
        "btn_buy"         : "🛒 Buy",
        "btn_sell"        : "💸 Sell",
        "select_asset_b"  : "💱 <b>Great choice! Which currency would you like to buy?</b>\n\nPick one below 👇",
        "select_asset_sl" : "💱 <b>Sure! Which currency would you like to sell?</b>\n\nPick one below 👇",
        "enter_amount"    : "💵 <b>How much would you like?</b>\n\nTap a quick amount below 👇 or press <b>Other</b> to enter a different amount.",
        "custom_amount"   : "✏️ <b>Enter the amount:</b>",
        "other_amount"    : "✍️ <b>Enter your desired amount:</b>\n\nType the number and send it 👇",
        "select_pay_cur"  : "🌍 <b>Which currency will you pay with?</b>\n\nSelect your payment currency below 👇",
        "enter_price"     : "💱 <b>What's your offer rate?</b>\n\nJust type the amount and send it 👇",
        "select_payment"  : "💳 <b>How would you like to pay?</b>\n\nChoose your preferred payment method 👇",
        "select_delivery" : "🚚 <b>How would you like to receive your funds?</b>\n\nChoose a delivery method below 👇",
        "enter_email"     : "📧 <b>Would you like a confirmation sent to your email?</b>\n\nEnter your email address or tap <b>Skip</b> if you'd rather not.",
        "btn_skip_email"  : "⏭️ Skip",
        "btn_rates"       : "💹 Exchange Rates",
        "btn_gold"        : "🥇 Gold Prices",
        "btn_back"        : "🏠 Back to Menu",
        "bad_amount"      : "⚠️ <b>That doesn't look right.</b>\n\nPlease enter a valid positive number and try again 👇",
        "bad_email"       : "⚠️ <b>That email doesn't look valid.</b>\n\nPlease try again or tap <b>Skip</b>.",
        "about": (
            "ℹ️ <b>About the Bot</b>\n\n"
            "Transfer Bot v4.0\n"
            "A fast and secure currency transfer platform.\n\n"
            "Available commands:\n"
            "/start — Begin a new transfer\n"
            "/about — Bot information\n\n"
            "Contact the admin for support."
        ),
        "type_lbl"     : "Type",
        "asset_lbl"    : "Asset",
        "amount_lbl"   : "Amount",
        "price_lbl"    : "Price",
        "payment_lbl"  : "Payment Method",
        "delivery_lbl" : "Delivery Method",
        "total_lbl"    : "Total",
        "rate_lbl"     : "Market Rate",
        "date_lbl"     : "Date",
        "done_title"   : "🎉 <b>Your order has been submitted!</b>\n\nThank you! Our team will be in touch with you shortly. Here's a summary of your order:",
    },
}

def t(lang: str, key: str) -> str:
    return TX.get(lang, TX["en"]).get(key, key)

# ══════════════════════════════════════════════
#  KEYBOARD HELPERS
# ══════════════════════════════════════════════
def _back_btn(lang: str) -> InlineKeyboardButton:
    return InlineKeyboardButton(t(lang, "btn_back"), callback_data="go_start")

def lang_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(row_width=2).add(
        InlineKeyboardButton("🇸🇦 العربية", callback_data="lang_ar"),
        InlineKeyboardButton("🇬🇧 English", callback_data="lang_en"),
    )

def type_kb(lang: str) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton(t(lang, "btn_buy"),  callback_data="type_buy"),
        InlineKeyboardButton(t(lang, "btn_sell"), callback_data="type_sell"),
    )
    return kb

def asset_kb(lang: str) -> InlineKeyboardMarkup:
    """Buy keyboard — shows all enabled currencies from admin settings."""
    currencies = get_enabled_currencies()
    kb = InlineKeyboardMarkup(row_width=3)
    kb.add(*[InlineKeyboardButton(c, callback_data=f"ast_{c}") for c in currencies])
    kb.add(_back_btn(lang))
    return kb

def sell_asset_kb(lang: str) -> InlineKeyboardMarkup:
    """Sell keyboard — always USDT and USDC only."""
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("💵 USDT", callback_data="ast_USDT"),
        InlineKeyboardButton("💵 USDC", callback_data="ast_USDC"),
    )
    kb.add(_back_btn(lang))
    return kb

def _back_to_asset_btn(lang: str) -> InlineKeyboardButton:
    label = "◀️ رجوع" if lang == "ar" else "◀️ Back"
    return InlineKeyboardButton(label, callback_data="back_to_asset")

def amount_kb(lang: str, asset: str) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=3)
    if asset in ("USD", "EUR", "USDT", "USDC"):
        kb.add(*[InlineKeyboardButton(str(a), callback_data=f"amt_{a}") for a in QUICK_AMOUNTS])
        other_label = "✏️ أخرى" if lang == "ar" else "✏️ Other"
        kb.add(InlineKeyboardButton(other_label, callback_data="amt_other"))
    kb.add(_back_to_asset_btn(lang))
    return kb

def _back_step_btn(lang: str, callback_data: str) -> InlineKeyboardButton:
    label = "◀️ رجوع" if lang == "ar" else "◀️ Back"
    return InlineKeyboardButton(label, callback_data=callback_data)

def pay_cur_kb(lang: str) -> InlineKeyboardMarkup:
    """Payment currency is always USD or EUR (not SYP)."""
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("🇺🇸 USD", callback_data="cur_USD"),
        InlineKeyboardButton("🇪🇺 EUR", callback_data="cur_EUR"),
    )
    kb.add(_back_step_btn(lang, "back_to_amount"))
    return kb

def price_back_kb(lang: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup().add(_back_step_btn(lang, "back_to_currency"))

def payment_method_kb(lang: str) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    for label, value in PAYMENT_METHODS.get(lang, PAYMENT_METHODS["en"]):
        kb.add(InlineKeyboardButton(label, callback_data=f"pay_{value}"))
    kb.add(_back_step_btn(lang, "back_to_calc"))
    return kb

def delivery_method_kb(lang: str) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    for label, value in DELIVERY_METHODS.get(lang, DELIVERY_METHODS["en"]):
        kb.add(InlineKeyboardButton(label, callback_data=f"del_{value}"))
    kb.add(_back_step_btn(lang, "back_to_payment"))
    return kb

def back_kb(lang: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup().add(_back_btn(lang))

_SECTION_LABELS = {
    "section_buy":    "🛒 Buy",
    "section_sell":   "💸 Sell",
    "section_rates":  "💹 Exchange Rates",
    "section_gold":   "🥇 Gold Prices",
    "section_crypto": "📊 Crypto Tracker",
    "USD": "💵 USD Currency",
    "EUR": "💶 EUR Currency",
    "SYP": "🇸🇾 SYP Currency",
}

def admin_kb(settings: list) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=1)
    settings_dict = dict(settings)

    # ── Sections group — same order as the keyboard ─────────
    SECTION_ORDER = [
        "section_buy",
        "section_sell",
        "section_rates",
        "section_gold",
        "section_crypto",
    ]
    kb.add(InlineKeyboardButton("━━━ 📂 Sections ━━━", callback_data="adm_noop"))
    for item in SECTION_ORDER:
        enabled = settings_dict.get(item, 1)
        icon  = "✅" if enabled else "🔴"
        label = _SECTION_LABELS.get(item, item)
        kb.add(InlineKeyboardButton(f"{icon}  {label}", callback_data=f"toggle_{item}"))

    # ── Currencies group ────────────────────────────────────
    kb.add(InlineKeyboardButton("━━━ 💰 Currencies ━━━", callback_data="adm_noop"))
    for item, enabled in settings:
        if not item.startswith("section_"):
            icon  = "✅" if enabled else "🔴"
            label = _SECTION_LABELS.get(item, item)
            kb.add(InlineKeyboardButton(f"{icon}  {label}", callback_data=f"toggle_{item}"))
    return kb

# ══════════════════════════════════════════════
#  WELCOME HELPER
# ══════════════════════════════════════════════
WELCOME_TEXT = (
    "🌟 <b>مرحبًا بك في بوت التحويل!</b> 🌟\n\n"
    "نسعى لجعل تجربة التحويل لديك سلسة، آمنة وذكية.\n\n"
    "💰 اختر العملة التي ترغب في استلامها\n"
    "💳 واختر طريقة الدفع المفضلة لديك\n\n"
    "للبدء، اضغط على /start واستعد لتجربة تحويل فريدة وسريعة! 🚀\n\n"
    "اختر لغتك للبدء 👇\n"
    "<i>Select your language to begin 👇</i>"
)

async def show_welcome(target, edit: bool = False) -> None:
    if edit:
        await target.edit_text(WELCOME_TEXT, reply_markup=lang_kb())
    else:
        await target.answer(WELCOME_TEXT, reply_markup=lang_kb())

# ══════════════════════════════════════════════
#  ── /start ──
# ══════════════════════════════════════════════
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove

PERSISTENT_KB = ReplyKeyboardMarkup(resize_keyboard=True, row_width=2).add(
    KeyboardButton("🛒 Buy / شراء"),
    KeyboardButton("💸 Sell / بيع"),
    KeyboardButton("💹 الأسعار / Rates"),
    KeyboardButton("🥇 الذهب / Gold"),
    KeyboardButton("📊 Crypto Tracker"),
)

class PortfolioStates(StatesGroup):
    entering_quantity = State()

class AdminTokenSearch(StatesGroup):
    waiting_query = State()


@dp.message_handler(commands=["start"], state="*")
async def cmd_start(message: types.Message, state: FSMContext):
    try:
        await state.finish()
        await message.answer("👇", reply_markup=PERSISTENT_KB)
        await show_welcome(message)
        await MarketFlow.language.set()
    except Exception as exc:
        logger.error("cmd_start: %s", exc)

# ══════════════════════════════════════════════
#  ── /rates shortcut ──
# ══════════════════════════════════════════════
async def show_rates_menu(target, lang: str, edit: bool = False):
    """Show the rates type selection menu (SYP or Global).
    target: a Message object.
    edit=True  → edit in place (use when coming back from a submenu).
    edit=False → send a new message (use when triggered from keyboard/command).
    """
    if lang == "ar":
        text = (
            "💹 <b>اختر نوع الأسعار</b>\n\n"
            "اختر القسم الذي تريد الاطلاع عليه 👇"
        )
        kb = InlineKeyboardMarkup(row_width=1).add(
            InlineKeyboardButton("🇸🇾 أسعار مقابل الليرة السورية", callback_data="syp_rates"),
            InlineKeyboardButton("🌍 أسعار عالمية  (SEK · USD · EUR)", callback_data="global_rates"),
            InlineKeyboardButton("🏠 القائمة الرئيسية", callback_data="back_to_main"),
        )
    else:
        text = (
            "💹 <b>Choose Rate Type</b>\n\n"
            "Select the section you'd like to view 👇"
        )
        kb = InlineKeyboardMarkup(row_width=1).add(
            InlineKeyboardButton("🇸🇾 Rates vs Syrian Pound (SYP)", callback_data="syp_rates"),
            InlineKeyboardButton("🌍 Global Rates  (SEK · USD · EUR)", callback_data="global_rates"),
            InlineKeyboardButton("🏠 Main Menu", callback_data="back_to_main"),
        )
    if edit:
        await target.edit_text(text, reply_markup=kb, parse_mode="HTML")
    else:
        await target.answer(text, reply_markup=kb, parse_mode="HTML")


@dp.message_handler(commands=["rates"], state="*")
async def cmd_rates(message: types.Message, state: FSMContext):
    try:
        if not await asyncio.to_thread(is_section_enabled, "section_rates"):
            await message.answer(SECTION_CLOSED_MSG["section_rates"])
            return
        data = await state.get_data()
        lang = data.get("language", "ar")
        await show_rates_menu(message, lang, edit=False)
    except Exception as exc:
        logger.error("cmd_rates: %s", exc)

# ══════════════════════════════════════════════
#  ── RATES: SYP section ──
#  FIX: get_sptoday_rates() is now run in a thread so it
#       doesn't block the event loop while fetching HTTP data.
# ══════════════════════════════════════════════
@dp.callback_query_handler(lambda c: c.data in ("syp_rates", "refresh_syp_rates"), state="*")
async def cb_syp_rates(callback: types.CallbackQuery, state: FSMContext):
    try:
        data = await state.get_data()
        lang = data.get("language", "ar")
        is_refresh = callback.data == "refresh_syp_rates"
        if is_refresh:
            await callback.answer("🔄 جارٍ التحديث..." if lang == "ar" else "🔄 Refreshing...", show_alert=False)
        else:
            loading = "⏳ جارٍ تحميل أسعار الليرة..." if lang == "ar" else "⏳ Loading SYP rates..."
            await callback.message.edit_text(loading)
        rates = await asyncio.to_thread(get_sptoday_rates)
        text  = format_rates_message(rates, lang)
        kb = InlineKeyboardMarkup(row_width=2)
        kb.add(
            InlineKeyboardButton("🔄 تحديث / Refresh", callback_data="refresh_syp_rates"),
        )
        kb.add(
            InlineKeyboardButton("◀️ رجوع / Back", callback_data="back_to_rates_menu"),
            InlineKeyboardButton("🏠 القائمة / Menu", callback_data="back_to_main"),
        )
        await callback.message.edit_text(text, reply_markup=kb, parse_mode="HTML", disable_web_page_preview=True)
        if not is_refresh:
            await callback.answer()
    except Exception as exc:
        logger.error("cb_syp_rates: %s", exc)

# ══════════════════════════════════════════════
#  ── RATES: Global section ──
#  FIX: get_global_rates() is now run in a thread.
# ══════════════════════════════════════════════
@dp.callback_query_handler(lambda c: c.data in ("global_rates", "refresh_global_rates"), state="*")
async def cb_global_rates(callback: types.CallbackQuery, state: FSMContext):
    try:
        data = await state.get_data()
        lang = data.get("language", "ar")
        is_refresh = callback.data == "refresh_global_rates"
        if is_refresh:
            await callback.answer("🔄 جارٍ التحديث..." if lang == "ar" else "🔄 Refreshing...", show_alert=False)
        else:
            loading = "⏳ جارٍ تحميل الأسعار العالمية..." if lang == "ar" else "⏳ Loading global rates..."
            await callback.message.edit_text(loading)
        pairs = await asyncio.to_thread(get_global_rates)
        text  = format_global_rates_message(pairs, lang)
        kb = InlineKeyboardMarkup(row_width=2)
        kb.add(
            InlineKeyboardButton("🔄 تحديث / Refresh", callback_data="refresh_global_rates"),
        )
        kb.add(
            InlineKeyboardButton("◀️ رجوع / Back", callback_data="back_to_rates_menu"),
            InlineKeyboardButton("🏠 القائمة / Menu", callback_data="back_to_main"),
        )
        await callback.message.edit_text(text, reply_markup=kb, parse_mode="HTML", disable_web_page_preview=True)
        if not is_refresh:
            await callback.answer()
    except Exception as exc:
        logger.error("cb_global_rates: %s", exc)

# ══════════════════════════════════════════════
#  ── RATES: Back to menu ──
# ══════════════════════════════════════════════
@dp.callback_query_handler(lambda c: c.data == "back_to_rates_menu", state="*")
async def cb_back_to_rates_menu(callback: types.CallbackQuery, state: FSMContext):
    try:
        data = await state.get_data()
        lang = data.get("language", "ar")
        await show_rates_menu(callback.message, lang, edit=True)
        await callback.answer()
    except Exception as exc:
        logger.error("cb_back_to_rates_menu: %s", exc)


@dp.message_handler(commands=["gold"], state="*")
async def cmd_gold(message: types.Message, state: FSMContext):
    try:
        if not await asyncio.to_thread(is_section_enabled, "section_gold"):
            await message.answer(SECTION_CLOSED_MSG["section_gold"])
            return
        data = await state.get_data()
        lang = data.get("language", "ar")
        loading = "⏳ جارٍ تحميل أسعار الذهب..." if lang == "ar" else "⏳ Loading gold prices..."
        msg = await message.answer(loading)
        gold = await asyncio.to_thread(get_gold_rates)
        text = format_gold_message(gold, lang)
        kb = InlineKeyboardMarkup(row_width=2)
        kb.add(InlineKeyboardButton("🔄 تحديث / Refresh", callback_data="refresh_gold"))
        kb.add(InlineKeyboardButton("🏠 القائمة / Menu", callback_data="back_to_main"))
        await msg.edit_text(text, reply_markup=kb, parse_mode="HTML", disable_web_page_preview=True)
    except Exception as exc:
        logger.error("cmd_gold: %s", exc)


@dp.callback_query_handler(lambda c: c.data == "refresh_gold", state="*")
async def cb_refresh_gold(callback: types.CallbackQuery, state: FSMContext):
    try:
        data = await state.get_data()
        lang = data.get("language", "ar")
        await callback.answer("🔄 جارٍ التحديث..." if lang == "ar" else "🔄 Refreshing...", show_alert=False)
        gold = await asyncio.to_thread(get_gold_rates)
        text = format_gold_message(gold, lang)
        kb = InlineKeyboardMarkup(row_width=2)
        kb.add(InlineKeyboardButton("🔄 تحديث / Refresh", callback_data="refresh_gold"))
        kb.add(InlineKeyboardButton("🏠 القائمة / Menu", callback_data="back_to_main"))
        await callback.message.edit_text(text, reply_markup=kb, parse_mode="HTML", disable_web_page_preview=True)
    except Exception as exc:
        logger.error("cb_refresh_gold: %s", exc)

# ══════════════════════════════════════════════
#  ── /help ──
# ══════════════════════════════════════════════
@dp.message_handler(commands=["help"], state="*")
async def cmd_help(message: types.Message, state: FSMContext):
    try:
        data = await state.get_data()
        lang = data.get("language", "ar")
        if lang == "ar":
            text = (
                "❓ <b>المساعدة</b>\n\n"
                "🏠 /start — القائمة الرئيسية\n"
                "💹 /rates — أسعار الصرف الحية\n"
                "🥇 /gold  — أسعار الذهب\n"
                "ℹ️ /about — عن البوت\n\n"
                "📊 /stats — إحصائيات (للمشرف فقط)\n"
                "⚙️ /admin — لوحة التحكم (للمشرف فقط)\n\n"
                "للبدء اضغط /start"
            )
        else:
            text = (
                "❓ <b>Help</b>\n\n"
                "🏠 /start — Main Menu\n"
                "💹 /rates — Live Exchange Rates\n"
                "🥇 /gold  — Live Gold Prices\n"
                "ℹ️ /about — About this bot\n\n"
                "📊 /stats — Statistics (Admin only)\n"
                "⚙️ /admin — Admin Panel (Admin only)\n\n"
                "Press /start to begin"
            )
        await message.answer(text)
    except Exception as exc:
        logger.error("cmd_help: %s", exc)


@dp.message_handler(commands=["about"], state="*")
async def cmd_about(message: types.Message, state: FSMContext):
    try:
        data = await state.get_data()
        lang = data.get("language", "ar")
        if lang == "ar":
            text = (
                "ℹ️ <b>عن بوت التحويل</b>\n\n"
                "الإصدار: <b>4.0</b>\n\n"
                "منصة آمنة وذكية لتحويل العملات الرقمية.\n\n"
                "✅ شراء وبيع USDT / USDC\n"
                "✅ أسعار صرف حية مقابل الليرة السورية\n"
                "✅ أسعار ذهب لحظية لجميع العيارات\n"
                "✅ أسعار عالمية (SEK · USD · EUR)\n"
                "✅ حساب تلقائي شفاف قبل التأكيد\n\n"
                "للدعم تواصل مع المشرف."
            )
        else:
            text = (
                "ℹ️ <b>About Transfer Bot</b>\n\n"
                "Version: <b>4.0</b>\n\n"
                "A secure, smart platform for crypto currency transfers.\n\n"
                "✅ Buy & Sell USDT / USDC\n"
                "✅ Live exchange rates vs Syrian Pound\n"
                "✅ Live gold prices for all karats\n"
                "✅ Global rates (SEK · USD · EUR)\n"
                "✅ Transparent calculation preview before confirming\n\n"
                "Contact the admin for support."
            )
        await message.answer(text)
    except Exception as exc:
        logger.error("cmd_about: %s", exc)

# ── Persistent keyboard button handlers ──
@dp.message_handler(lambda m: m.text == "🛒 Buy / شراء", state="*")
async def kb_buy_direct(message: types.Message, state: FSMContext):
    """Buy button — ask language then jump straight to asset selection."""
    try:
        if not await asyncio.to_thread(is_section_enabled, "section_buy"):
            await message.answer(SECTION_CLOSED_MSG["section_buy"])
            return
        await state.finish()
        await state.update_data(preset_type="buy")
        kb = InlineKeyboardMarkup().add(
            InlineKeyboardButton("🇸🇦 العربية", callback_data="dlang_ar"),
            InlineKeyboardButton("🇬🇧 English", callback_data="dlang_en"),
        )
        await message.answer("🛒 اختر اللغة / Choose language:", reply_markup=kb)
        await MarketFlow.language.set()
    except Exception as exc:
        logger.error("kb_buy_direct: %s", exc)

@dp.message_handler(lambda m: m.text == "💸 Sell / بيع", state="*")
async def kb_sell_direct(message: types.Message, state: FSMContext):
    """Sell button — ask language then jump straight to asset selection."""
    try:
        if not await asyncio.to_thread(is_section_enabled, "section_sell"):
            await message.answer(SECTION_CLOSED_MSG["section_sell"])
            return
        await state.finish()
        await state.update_data(preset_type="sell")
        kb = InlineKeyboardMarkup().add(
            InlineKeyboardButton("🇸🇦 العربية", callback_data="dlang_ar"),
            InlineKeyboardButton("🇬🇧 English", callback_data="dlang_en"),
        )
        await message.answer("💸 اختر اللغة / Choose language:", reply_markup=kb)
        await MarketFlow.language.set()
    except Exception as exc:
        logger.error("kb_sell_direct: %s", exc)

@dp.message_handler(lambda m: m.text in ["🥇 الذهب / Gold"], state="*")
async def kb_gold(message: types.Message, state: FSMContext):
    if not await asyncio.to_thread(is_section_enabled, "section_gold"):
        await message.answer(SECTION_CLOSED_MSG["section_gold"])
        return
    await cmd_gold(message, state)

@dp.message_handler(lambda m: m.text == "📊 Crypto Tracker", state="*")
async def kb_crypto_tracker(message: types.Message, state: FSMContext):
    """Handle Crypto Tracker keyboard button press."""
    if not await asyncio.to_thread(is_section_enabled, "section_crypto"):
        await message.answer(SECTION_CLOSED_MSG["section_crypto"])
        return
    await cmd_tracker(message)

@dp.message_handler(lambda m: m.text in ["💹 الأسعار / Rates"], state="*")
async def kb_rates(message: types.Message, state: FSMContext):
    if not await asyncio.to_thread(is_section_enabled, "section_rates"):
        await message.answer(SECTION_CLOSED_MSG["section_rates"])
        return
    await cmd_rates(message, state)


# ══════════════════════════════════════════════
#  ── /stats  (admin only) ──
# ══════════════════════════════════════════════
@dp.message_handler(commands=["stats"], state="*")
async def cmd_stats(message: types.Message):
    try:
        if message.from_user.id != ADMIN_ID:
            await message.answer("🚫 Unauthorised.")
            return
        s = get_stats()
        vol_lines = "\n".join(
            f"  ├ {asset}: <b>{vol:,.2f}</b>" for asset, vol in s["volumes"]
        ) or "  └ No trades yet"
        text = (
            "📊 <b>إحصائيات / Trade Statistics</b>\n"
            "━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"📋 الإجمالي / Total  : <b>{s['total']}</b>\n"
            f"  ├ 🛒 شراء / Buy    : <b>{s['buy']}</b>\n"
            f"  └ 💸 بيع / Sell    : <b>{s['sell']}</b>\n\n"
            f"🔄 الحالة / Status\n"
            f"  ├ ⏳ Pending        : <b>{s['pending']}</b>\n"
            f"  └ ✅ Completed      : <b>{s['completed']}</b>\n\n"
            f"💰 الحجم / Volume by Asset\n"
            f"{vol_lines}"
        )
        await message.answer(text)
    except Exception as exc:
        logger.error("cmd_stats: %s", exc)

# ══════════════════════════════════════════════
#  ── /admin ──
# ══════════════════════════════════════════════
@dp.message_handler(commands=["admin"], state="*")
async def cmd_admin(message: types.Message):
    try:
        if message.from_user.id != ADMIN_ID:
            await message.answer("🚫 Unauthorised.")
            return
        settings = get_all_settings()
        await message.answer(
            "⚙️ <b>Admin Panel</b>\nTap a currency to toggle it on/off:",
            reply_markup=admin_kb(settings),
        )
    except Exception as exc:
        logger.error("cmd_admin: %s", exc)

@dp.callback_query_handler(lambda c: c.data == "adm_noop", state="*")
async def cb_adm_noop(callback: types.CallbackQuery):
    await callback.answer()

@dp.callback_query_handler(lambda c: c.data.startswith("toggle_"), state="*")
async def cb_toggle(callback: types.CallbackQuery):
    try:
        if callback.from_user.id != ADMIN_ID:
            await callback.answer("🚫 Unauthorised", show_alert=True)
            return
        item    = callback.data.split("_", 1)[1]
        # rebuild the real key (section items are stored as "section_X")
        new_val = toggle_setting(item)
        label   = _SECTION_LABELS.get(item, item)
        status  = "✅ Enabled" if new_val else "🔴 Disabled"
        await callback.message.edit_reply_markup(reply_markup=admin_kb(get_all_settings()))
        await callback.answer(f"{label} — {status}", show_alert=True)
    except Exception as exc:
        logger.error("cb_toggle: %s", exc)

# ══════════════════════════════════════════════
#  ── LANGUAGE ──
# ══════════════════════════════════════════════
@dp.callback_query_handler(lambda c: c.data.startswith("dlang_"), state=MarketFlow.language)
async def cb_direct_language(callback: types.CallbackQuery, state: FSMContext):
    """Language selection from direct Buy/Sell buttons — skips type step."""
    try:
        lang = callback.data.split("_", 1)[1]
        await state.update_data(language=lang)
        await asyncio.to_thread(save_user_lang, callback.from_user.id, lang)
        data = await state.get_data()
        trade_type = data.get("preset_type", "buy")
        await state.update_data(type=trade_type)

        if trade_type == "buy":
            await callback.message.edit_text(t(lang, "select_asset_b"), reply_markup=asset_kb(lang))
        else:
            await callback.message.edit_text(t(lang, "select_asset_sl"), reply_markup=sell_asset_kb(lang))

        await callback.answer()
        await MarketFlow.asset.set()
    except Exception as exc:
        logger.error("cb_direct_language: %s", exc)

@dp.callback_query_handler(lambda c: c.data.startswith("lang_"), state=MarketFlow.language)
async def cb_language(callback: types.CallbackQuery, state: FSMContext):
    try:
        lang = callback.data.split("_", 1)[1]
        await state.update_data(language=lang)
        await asyncio.to_thread(save_user_lang, callback.from_user.id, lang)
        await callback.message.edit_text(t(lang, "select_type"), reply_markup=type_kb(lang))
        await callback.answer()
        await MarketFlow.type.set()
    except Exception as exc:
        logger.error("cb_language: %s", exc)

# ══════════════════════════════════════════════
#  ── TYPE  (Buy / Sell) ──
#  FIX: Buy now shows asset_kb (admin-controlled currencies),
#       Sell shows sell_asset_kb (USDT/USDC only).
# ══════════════════════════════════════════════
@dp.callback_query_handler(lambda c: c.data.startswith("type_"), state=MarketFlow.type)
async def cb_type(callback: types.CallbackQuery, state: FSMContext):
    try:
        data       = await state.get_data()
        lang       = data.get("language", "en")
        trade_type = callback.data.split("_", 1)[1]
        await state.update_data(type=trade_type)

        asset_key_map = {
            "buy":  "select_asset_b",
            "sell": "select_asset_sl",
        }
        asset_key = asset_key_map.get(trade_type, "select_asset_b")

        # FIX: use the correct keyboard for buy vs sell
        kb = asset_kb(lang) if trade_type == "buy" else sell_asset_kb(lang)

        await callback.message.edit_text(t(lang, asset_key), reply_markup=kb)
        await callback.answer()
        await MarketFlow.asset.set()
    except Exception as exc:
        logger.error("cb_type: %s", exc)

# ══════════════════════════════════════════════
#  ── ASSET ──
# ══════════════════════════════════════════════
@dp.callback_query_handler(lambda c: c.data.startswith("ast_"), state=MarketFlow.asset)
async def cb_asset(callback: types.CallbackQuery, state: FSMContext):
    try:
        data  = await state.get_data()
        lang  = data.get("language", "en")
        asset = callback.data.split("_", 1)[1]
        await state.update_data(asset=asset)

        text = t(lang, "enter_amount")
        if asset not in ("USD", "EUR"):
            text += f"\n{t(lang, 'custom_amount')}"
        await callback.message.edit_text(text, reply_markup=amount_kb(lang, asset))
        await callback.answer()
        await MarketFlow.amount.set()
    except Exception as exc:
        logger.error("cb_asset: %s", exc)

# ══════════════════════════════════════════════
#  ── AMOUNT  (quick button) ──
# ══════════════════════════════════════════════
@dp.callback_query_handler(lambda c: c.data.startswith("amt_"), state=MarketFlow.amount)
async def cb_amount_quick(callback: types.CallbackQuery, state: FSMContext):
    try:
        data    = await state.get_data()
        lang    = data.get("language", "en")
        raw     = callback.data.split("_", 1)[1]

        if raw == "other":
            # Ask user to type a custom amount
            kb = InlineKeyboardMarkup().add(_back_to_asset_btn(lang))
            await callback.message.edit_text(
                t(lang, "other_amount"),
                reply_markup=kb,
                parse_mode="HTML"
            )
            # Stay in MarketFlow.amount state so msg_amount picks it up
            await callback.answer()
            return

        amount = float(raw)
        await state.update_data(amount=amount)
        await callback.answer()
        await callback.message.edit_text(t(lang, "select_pay_cur"), reply_markup=pay_cur_kb(lang), parse_mode="HTML")
        await MarketFlow.currency.set()
    except Exception as exc:
        logger.error("cb_amount_quick: %s", exc)

# ══════════════════════════════════════════════
#  ── AMOUNT  (typed) ──
# ══════════════════════════════════════════════
@dp.message_handler(state=MarketFlow.amount)
async def msg_amount(message: types.Message, state: FSMContext):
    try:
        if is_rate_limited(message.from_user.id):
            return
        data = await state.get_data()
        lang = data.get("language", "en")
        try:
            amount = float(message.text.replace(",", "").strip())
            assert amount > 0
        except (ValueError, AssertionError):
            await message.answer(t(lang, "bad_amount"))
            return
        await state.update_data(amount=amount)
        await message.answer(t(lang, "select_pay_cur"), reply_markup=pay_cur_kb(lang))
        await MarketFlow.currency.set()
    except Exception as exc:
        logger.error("msg_amount: %s", exc)

# ══════════════════════════════════════════════
#  ── PAYMENT CURRENCY ──
#  FIX: fetch_crypto_reference runs in a thread
# ══════════════════════════════════════════════
@dp.callback_query_handler(lambda c: c.data.startswith("cur_"), state=MarketFlow.currency)
async def cb_currency(callback: types.CallbackQuery, state: FSMContext):
    try:
        data     = await state.get_data()
        lang     = data.get("language", "en")
        currency = callback.data.split("_", 1)[1]
        await state.update_data(currency=currency)
        asset = data.get("asset", "USDT")

        live_rate, rate_source = await asyncio.to_thread(fetch_crypto_reference, asset, currency)
        price_msg = _build_price_entry_msg(lang, asset, currency, live_rate, rate_source)

        await callback.message.edit_text(price_msg, parse_mode="HTML", reply_markup=price_back_kb(lang))
        await callback.answer()
        await MarketFlow.price.set()
    except Exception as exc:
        logger.error("cb_currency: %s", exc)

# ══════════════════════════════════════════════
#  ── PRICE ──
# ══════════════════════════════════════════════
@dp.message_handler(state=MarketFlow.price)
async def msg_price(message: types.Message, state: FSMContext):
    try:
        if is_rate_limited(message.from_user.id):
            return
        data = await state.get_data()
        lang = data.get("language", "en")
        try:
            price = float(message.text.replace(",", "").strip())
            assert price > 0
        except (ValueError, AssertionError):
            await message.answer(t(lang, "bad_amount"))
            return
        await state.update_data(price=price)

        # Build calculation preview
        amount   = data.get("amount", 0)
        asset    = data.get("asset", "USDT")
        currency = data.get("currency", "USD")

        preview = _build_calc_preview(lang, amount, asset, price, currency)
        await message.answer(preview, parse_mode="HTML", reply_markup=_calc_confirm_kb(lang))
        await MarketFlow.confirm_calc.set()
    except Exception as exc:
        logger.error("msg_price: %s", exc)


# ══════════════════════════════════════════════
#  ── CALCULATION CONFIRM / EDIT ──
#  FIX: fetch_crypto_reference runs in a thread
# ══════════════════════════════════════════════
@dp.callback_query_handler(lambda c: c.data in ("calc_confirm", "calc_edit"), state=MarketFlow.confirm_calc)
async def cb_calc_confirm(callback: types.CallbackQuery, state: FSMContext):
    try:
        data = await state.get_data()
        lang = data.get("language", "en")
        await callback.answer()

        if callback.data == "calc_edit":
            # Go back to price entry — re-fetch live rate
            asset    = data.get("asset", "USDT")
            currency = data.get("currency", "USD")
            live_rate, rate_source = await asyncio.to_thread(fetch_crypto_reference, asset, currency)
            price_msg = _build_price_entry_msg(lang, asset, currency, live_rate, rate_source)
            await callback.message.edit_text(price_msg, parse_mode="HTML", reply_markup=price_back_kb(lang))
            await MarketFlow.price.set()
        else:
            # Confirmed — proceed to payment method
            await callback.message.edit_text(t(lang, "select_payment"), reply_markup=payment_method_kb(lang))
            await MarketFlow.payment.set()
    except Exception as exc:
        logger.error("cb_calc_confirm: %s", exc)

# ══════════════════════════════════════════════
#  ── PAYMENT METHOD ──
# ══════════════════════════════════════════════
@dp.callback_query_handler(lambda c: c.data.startswith("pay_"), state=MarketFlow.payment)
async def cb_payment(callback: types.CallbackQuery, state: FSMContext):
    try:
        data    = await state.get_data()
        lang    = data.get("language", "en")
        value   = callback.data.split("_", 1)[1]
        methods = PAYMENT_METHODS.get(lang, PAYMENT_METHODS["en"])
        label   = next((lbl for lbl, val in methods if val == value), value)
        await state.update_data(payment=label)
        await callback.message.edit_text(t(lang, "select_delivery"), reply_markup=delivery_method_kb(lang))
        await callback.answer()
        await MarketFlow.delivery.set()
    except Exception as exc:
        logger.error("cb_payment: %s", exc)

# ══════════════════════════════════════════════
#  ── DELIVERY METHOD ──
# ══════════════════════════════════════════════
@dp.callback_query_handler(lambda c: c.data.startswith("del_"), state=MarketFlow.delivery)
async def cb_delivery(callback: types.CallbackQuery, state: FSMContext):
    try:
        data    = await state.get_data()
        lang    = data.get("language", "en")
        value   = callback.data.split("_", 1)[1]
        methods = DELIVERY_METHODS.get(lang, DELIVERY_METHODS["en"])
        label   = next((lbl for lbl, val in methods if val == value), value)
        await state.update_data(delivery=label)
        email_kb = InlineKeyboardMarkup(row_width=2).add(
            InlineKeyboardButton(t(lang, "btn_skip_email"), callback_data="skip_email"),
            _back_step_btn(lang, "back_to_delivery"),
        )
        await callback.message.edit_text(t(lang, "enter_email"), reply_markup=email_kb)
        await callback.answer()
        await MarketFlow.email.set()
    except Exception as exc:
        logger.error("cb_delivery: %s", exc)

# ══════════════════════════════════════════════
#  ── EMAIL + FINAL SUMMARY ──
#  FIX: save_trade and send_confirmation_email
#       both run in threads (non-blocking).
# ══════════════════════════════════════════════
async def _finish_trade(lang: str, data: dict, user, send_to: Optional[str], reply_target):
    """Shared logic: save trade, build summary, post to group, send email if provided."""
    row       = await asyncio.to_thread(save_trade, data, user)
    total     = row[11]
    eur_snap  = row[13]
    rate_line = f"1 USD = {eur_snap:.4f} EUR" if eur_snap else "N/A"

    asset    = data.get("asset", "USDT")
    currency = data.get("currency", "USD")
    amount   = data.get("amount", 0)
    price    = data.get("price", 0)

    sep = "━" * 30
    mid = "─" * 30
    if lang == "ar":
        summary = (
            f"{t(lang, 'done_title')}\n\n"
            f"🆔 <b>طلب رقم #{row[0]:05d}</b>\n"
            f"{sep}\n"
            f"📌 <b>النوع:</b>          <b>{'شراء' if data['type']=='buy' else 'بيع'} {asset}</b>\n"
            f"📦 <b>الكمية:</b>         <b>{amount:,.2f} {asset}</b>\n"
            f"{mid}\n"
            f"💱 <b>السعر المطلوب:</b>  <b>{price:,.2f} {currency}</b> لكل 100 {asset}\n"
            f"🧾 <b>الحساب:</b>         ({amount:,.2f} ÷ 100) × {price:,.2f}\n"
            f"{mid}\n"
            f"💰 <b>الإجمالي:</b>       <b>{total:,.2f} {currency}</b>\n"
            f"{sep}\n"
            f"💳 <b>طريقة الدفع:</b>   {data.get('payment', '—')}\n"
            f"🚚 <b>طريقة الاستلام:</b> {data.get('delivery', '—')}\n"
            f"📅 <b>التاريخ:</b>        {row[15]}\n"
        )
    else:
        summary = (
            f"{t(lang, 'done_title')}\n\n"
            f"🆔 <b>Order #{row[0]:05d}</b>\n"
            f"{sep}\n"
            f"📌 <b>Type:</b>        <b>{'BUY' if data['type']=='buy' else 'SELL'} {asset}</b>\n"
            f"📦 <b>Amount:</b>      <b>{amount:,.2f} {asset}</b>\n"
            f"{mid}\n"
            f"💱 <b>Your Rate:</b>   <b>{price:,.2f} {currency}</b> per 100 {asset}\n"
            f"🧾 <b>Calculation:</b> ({amount:,.2f} ÷ 100) × {price:,.2f}\n"
            f"{mid}\n"
            f"💰 <b>Total:</b>       <b>{total:,.2f} {currency}</b>\n"
            f"{sep}\n"
            f"💳 <b>Payment:</b>     {data.get('payment', '—')}\n"
            f"🚚 <b>Delivery:</b>    {data.get('delivery', '—')}\n"
            f"📅 <b>Date:</b>        {row[15]}\n"
        )

    if hasattr(reply_target, 'message'):
        await reply_target.message.edit_text(summary, reply_markup=back_kb(lang), parse_mode="HTML")
    else:
        await reply_target.answer(summary, reply_markup=back_kb(lang), parse_mode="HTML")

    # ── Professional group notification ──
    try:
        user_mention = (
            f"@{user.username}" if user.username
            else f"<a href='tg://user?id={user.id}'>{user.full_name}</a>"
        )
        trade_type_label = (
            ("🛒 BUY"  if data["type"] == "buy" else "💸 SELL")
        )
        group_text = (
            f"🔔 <b>New Order  #{row[0]:05d}</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"👤 <b>User:</b>      {user_mention}  (<code>{user.id}</code>)\n"
            f"📌 <b>Type:</b>      <b>{trade_type_label} {asset}</b>\n"
            f"📦 <b>Amount:</b>    <code>{amount:,.2f} {asset}</code>\n"
            f"💱 <b>Rate:</b>      <code>{price:,.2f} {currency}</code> per 100 {asset}\n"
            f"💰 <b>Total:</b>     <b>{total:,.2f} {currency}</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"💳 <b>Payment:</b>   {data.get('payment', '—')}\n"
            f"🚚 <b>Delivery:</b>  {data.get('delivery', '—')}\n"
            f"📧 <b>Email:</b>     {data.get('email') or '—'}\n"
            f"📅 <b>Date:</b>      {row[15]}\n"
            f"🔄 <b>Status:</b>    ⏳ PENDING"
        )
        await bot.send_message(GROUP_ID, group_text, parse_mode="HTML")
    except Exception as ge:
        logger.warning("Group post failed: %s", ge)

    # FIX: email is sent in a background thread — never blocks the bot
    if send_to:
        asyncio.create_task(asyncio.to_thread(send_confirmation_email, send_to, row))


@dp.callback_query_handler(lambda c: c.data == "skip_email", state=MarketFlow.email)
async def cb_skip_email(callback: types.CallbackQuery, state: FSMContext):
    try:
        data = await state.get_data()
        lang = data.get("language", "en")
        await state.update_data(email="")
        data = await state.get_data()
        await _finish_trade(lang, data, callback.from_user, None, callback)
        await state.finish()
    except Exception as exc:
        logger.error("cb_skip_email: %s", exc)


@dp.message_handler(state=MarketFlow.email)
async def msg_email(message: types.Message, state: FSMContext):
    try:
        if is_rate_limited(message.from_user.id):
            return
        data  = await state.get_data()
        lang  = data.get("language", "en")
        email = message.text.strip()

        if not EMAIL_RE.match(email):
            email_kb = InlineKeyboardMarkup(row_width=2).add(
                InlineKeyboardButton(t(lang, "btn_skip_email"), callback_data="skip_email"),
                InlineKeyboardButton(t(lang, "btn_back"),       callback_data="back_to_delivery"),
            )
            await message.answer(t(lang, "bad_email"), reply_markup=email_kb)
            return

        await state.update_data(email=email)
        data = await state.get_data()
        await _finish_trade(lang, data, message.from_user, email, message)
        await state.finish()

    except Exception as exc:
        logger.error("msg_email: %s", exc)

# ══════════════════════════════════════════════
#  ── COINGECKO LIVE RATE HELPER ──
# ══════════════════════════════════════════════
def _build_price_entry_msg(lang: str, asset: str, currency: str,
                           live_rate, rate_source) -> str:
    """Build the bilingual price-entry prompt (used in 3 places)."""
    cur_symbol = "€" if currency == "EUR" else "$"
    if live_rate is not None:
        rate_line_en = (
            f"\n📊 <b>Live Rate:</b>  1 {asset} ≈ <b>{cur_symbol}{live_rate:.4f}</b>"
            f"  <i>({rate_source})</i>"
        )
        rate_line_ar = (
            f"\n📊 <b>السعر الحالي:</b>  1 {asset} ≈ <b>{cur_symbol}{live_rate:.4f}</b>"
            f"  <i>({rate_source})</i>"
        )
    else:
        rate_line_en = "\n⚠️ <i>Live rate unavailable right now.</i>"
        rate_line_ar = "\n⚠️ <i>تعذّر جلب السعر الحالي في الوقت الفعلي.</i>"

    if lang == "ar":
        return (
            f"💱 <b>ما هو سعرك المطلوب؟</b>\n\n"
            f"أدخل كم <b>{currency}</b> تريد مقابل كل <b>100 {asset}</b>\n"
            f"✏️ مثال: اكتب <code>101</code> إذا كان سعرك <b>101 {currency}</b>"
            f" لكل <b>100 {asset}</b>\n"
            f"{rate_line_ar}"
            f"\n\nاكتب المبلغ وأرسله 👇"
        )
    else:
        return (
            f"💱 <b>What's your offer rate?</b>\n\n"
            f"Enter how much <b>{currency}</b> per <b>100 {asset}</b>\n"
            f"✏️ Example: type <code>101</code> if your rate is <b>101 {currency}</b>"
            f" per <b>100 {asset}</b>\n"
            f"{rate_line_en}"
            f"\n\nJust type the amount and send it 👇"
        )


def _build_calc_preview(lang: str, amount: float, asset: str,
                        price: float, currency: str) -> str:
    """Build the bilingual calculation-preview message."""
    total = (amount / 100) * price
    sep   = "━" * 30
    mid   = "─" * 30
    if lang == "ar":
        return (
            f"🧮 <b>ملخص الحساب</b>\n"
            f"{sep}\n"
            f"📦 <b>الكمية:</b>          <code>{amount:,.2f} {asset}</code>\n"
            f"💱 <b>سعرك المطلوب:</b>    <code>{price:,.2f} {currency}</code>"
            f" لكل 100 {asset}\n"
            f"{mid}\n"
            f"🧾 <b>الحساب:</b>  ({amount:,.2f} ÷ 100) × {price:,.2f}"
            f" = <b>{total:,.2f} {currency}</b>\n"
            f"{sep}\n"
            f"💰 <b>الإجمالي المتوقع:  {total:,.2f} {currency}</b>\n\n"
            f"هل الحساب صحيح؟ اضغط تأكيد للمتابعة 👇"
        )
    else:
        return (
            f"🧮 <b>Calculation Summary</b>\n"
            f"{sep}\n"
            f"📦 <b>Amount:</b>         <code>{amount:,.2f} {asset}</code>\n"
            f"💱 <b>Your Rate:</b>      <code>{price:,.2f} {currency}</code>"
            f" per 100 {asset}\n"
            f"{mid}\n"
            f"🧾 <b>Calculation:</b>  ({amount:,.2f} ÷ 100) × {price:,.2f}"
            f" = <b>{total:,.2f} {currency}</b>\n"
            f"{sep}\n"
            f"💰 <b>Total:  {total:,.2f} {currency}</b>\n\n"
            f"Does everything look correct? Tap Confirm to continue 👇"
        )


def _calc_confirm_kb(lang: str) -> InlineKeyboardMarkup:
    """Keyboard for the calculation preview screen."""
    return InlineKeyboardMarkup(row_width=2).add(
        InlineKeyboardButton(
            "✅ تأكيد" if lang == "ar" else "✅ Confirm",
            callback_data="calc_confirm"
        ),
        InlineKeyboardButton(
            "✏️ تعديل السعر" if lang == "ar" else "✏️ Edit Rate",
            callback_data="calc_edit"
        ),
    )


def fetch_crypto_reference(asset: str, currency: str):
    """
    Fetches live crypto rate for the price entry screen reference.
    Tries CoinGecko first, falls back to CryptoCompare if it fails.
    asset: USDT or USDC  |  currency: USD or EUR
    Returns (price: float, source: str) or (None, None) on total failure.
    """
    # ── Source 1: CoinGecko ──
    try:
        coin_id = "tether" if asset == "USDT" else "usd-coin"
        vs = currency.lower()
        url = f"https://api.coingecko.com/api/v3/simple/price?ids={coin_id}&vs_currencies={vs}"
        resp = requests.get(url, timeout=6)
        resp.raise_for_status()
        price = resp.json()[coin_id][vs]
        return price, "CoinGecko"
    except Exception:
        pass

    # ── Source 2: CryptoCompare (fallback) ──
    try:
        url = f"https://min-api.cryptocompare.com/data/price?fsym={asset}&tsyms={currency}"
        resp = requests.get(url, timeout=6)
        resp.raise_for_status()
        price = resp.json()[currency]
        return price, "CryptoCompare"
    except Exception:
        pass

    return None, None

# ══════════════════════════════════════════════
#  ── EXCHANGE RATES (sp-today.com) ──
# ══════════════════════════════════════════════
@dp.callback_query_handler(lambda c: c.data == "show_gold", state="*")
async def cb_show_gold(callback: types.CallbackQuery, state: FSMContext):
    try:
        data = await state.get_data()
        lang = data.get("language", "ar")
        loading = "⏳ جارٍ تحميل أسعار الذهب..." if lang == "ar" else "⏳ Loading gold prices..."
        await callback.answer(loading)
        gold = await asyncio.to_thread(get_gold_rates)
        text = format_gold_message(gold, lang)
        kb = InlineKeyboardMarkup().add(
            InlineKeyboardButton(t(lang, "btn_back"), callback_data="go_start")
        )
        await callback.message.edit_text(text, reply_markup=kb, parse_mode="HTML", disable_web_page_preview=True)
    except Exception as exc:
        logger.error("cb_show_gold: %s", exc)


@dp.callback_query_handler(lambda c: c.data == "show_rates", state="*")
async def cb_show_rates(callback: types.CallbackQuery, state: FSMContext):
    try:
        data = await state.get_data()
        lang = data.get("language", "ar")
        await show_rates_menu(callback.message, lang, edit=True)
        await callback.answer()
    except Exception as exc:
        logger.error("cb_show_rates: %s", exc)


# ══════════════════════════════════════════════
#  ── BACK: CURRENCY → AMOUNT ──
# ══════════════════════════════════════════════
@dp.callback_query_handler(lambda c: c.data == "back_to_amount", state="*")
async def cb_back_to_amount(callback: types.CallbackQuery, state: FSMContext):
    try:
        data  = await state.get_data()
        lang  = data.get("language", "en")
        asset = data.get("asset", "USDT")
        text  = t(lang, "enter_amount")
        if asset not in ("USD", "EUR"):
            text += f"\n{t(lang, 'custom_amount')}"
        await callback.message.edit_text(text, reply_markup=amount_kb(lang, asset))
        await MarketFlow.amount.set()
        await callback.answer()
    except Exception as exc:
        logger.error("cb_back_to_amount: %s", exc)

# ══════════════════════════════════════════════
#  ── BACK: PRICE → CURRENCY ──
# ══════════════════════════════════════════════
@dp.callback_query_handler(lambda c: c.data == "back_to_currency", state="*")
async def cb_back_to_currency(callback: types.CallbackQuery, state: FSMContext):
    try:
        data = await state.get_data()
        lang = data.get("language", "en")
        await callback.message.edit_text(t(lang, "select_pay_cur"), reply_markup=pay_cur_kb(lang), parse_mode="HTML")
        await MarketFlow.currency.set()
        await callback.answer()
    except Exception as exc:
        logger.error("cb_back_to_currency: %s", exc)

# ══════════════════════════════════════════════
#  ── BACK: PAYMENT → PRICE ──
#  FIX: fetch_crypto_reference runs in a thread
# ══════════════════════════════════════════════
@dp.callback_query_handler(lambda c: c.data == "back_to_price", state="*")
async def cb_back_to_price(callback: types.CallbackQuery, state: FSMContext):
    try:
        data     = await state.get_data()
        lang     = data.get("language", "en")
        asset    = data.get("asset", "USDT")
        currency = data.get("currency", "USD")
        live_rate, rate_source = await asyncio.to_thread(fetch_crypto_reference, asset, currency)
        price_msg = _build_price_entry_msg(lang, asset, currency, live_rate, rate_source)
        await callback.message.edit_text(price_msg, parse_mode="HTML", reply_markup=price_back_kb(lang))
        await MarketFlow.price.set()
        await callback.answer()
    except Exception as exc:
        logger.error("cb_back_to_price: %s", exc)

# ══════════════════════════════════════════════
#  ── BACK: PAYMENT → CALC PREVIEW ──
# ══════════════════════════════════════════════
@dp.callback_query_handler(lambda c: c.data == "back_to_calc", state="*")
async def cb_back_to_calc(callback: types.CallbackQuery, state: FSMContext):
    try:
        data     = await state.get_data()
        lang     = data.get("language", "en")
        amount   = data.get("amount", 0)
        asset    = data.get("asset", "USDT")
        price    = data.get("price", 0)
        currency = data.get("currency", "USD")
        preview  = _build_calc_preview(lang, amount, asset, price, currency)
        await callback.message.edit_text(preview, parse_mode="HTML", reply_markup=_calc_confirm_kb(lang))
        await MarketFlow.confirm_calc.set()
        await callback.answer()
    except Exception as exc:
        logger.error("cb_back_to_calc: %s", exc)

# ══════════════════════════════════════════════
#  ── BACK: DELIVERY → PAYMENT ──
# ══════════════════════════════════════════════
@dp.callback_query_handler(lambda c: c.data == "back_to_payment", state="*")
async def cb_back_to_payment(callback: types.CallbackQuery, state: FSMContext):
    try:
        data = await state.get_data()
        lang = data.get("language", "en")
        await callback.message.edit_text(t(lang, "select_payment"), reply_markup=payment_method_kb(lang), parse_mode="HTML")
        await MarketFlow.payment.set()
        await callback.answer()
    except Exception as exc:
        logger.error("cb_back_to_payment: %s", exc)

# ══════════════════════════════════════════════
#  ── BACK: EMAIL → DELIVERY ──
# ══════════════════════════════════════════════
@dp.callback_query_handler(lambda c: c.data == "back_to_delivery", state="*")
async def cb_back_to_delivery(callback: types.CallbackQuery, state: FSMContext):
    try:
        data = await state.get_data()
        lang = data.get("language", "en")
        await callback.message.edit_text(t(lang, "select_delivery"), reply_markup=delivery_method_kb(lang), parse_mode="HTML")
        await MarketFlow.delivery.set()
        await callback.answer()
    except Exception as exc:
        logger.error("cb_back_to_delivery: %s", exc)

# ══════════════════════════════════════════════
#  ── BACK TO ASSET SELECTION ──
#  FIX: uses correct keyboard depending on buy/sell
# ══════════════════════════════════════════════
@dp.callback_query_handler(lambda c: c.data == "back_to_asset", state="*")
async def cb_back_to_asset(callback: types.CallbackQuery, state: FSMContext):
    try:
        data       = await state.get_data()
        lang       = data.get("language", "en")
        trade_type = data.get("type", "buy")
        asset_key_map = {
            "buy":  "select_asset_b",
            "sell": "select_asset_sl",
        }
        asset_key = asset_key_map.get(trade_type, "select_asset_b")
        # FIX: correct keyboard for buy vs sell
        kb = asset_kb(lang) if trade_type == "buy" else sell_asset_kb(lang)
        await callback.message.edit_text(t(lang, asset_key), reply_markup=kb)
        await MarketFlow.asset.set()
        await callback.answer()
    except Exception as exc:
        logger.error("cb_back_to_asset: %s", exc)

# ══════════════════════════════════════════════
#  ── BACK TO START ──
# ══════════════════════════════════════════════
@dp.callback_query_handler(lambda c: c.data in ("go_start", "back_to_main"), state="*")
async def cb_go_start(callback: types.CallbackQuery, state: FSMContext):
    try:
        await state.finish()
        await show_welcome(callback.message, edit=True)
        await MarketFlow.language.set()
        await callback.answer()
    except Exception as exc:
        logger.error("cb_go_start: %s", exc)

# ══════════════════════════════════════════════
#  ── UNKNOWN COMMANDS ──
# ══════════════════════════════════════════════
@dp.message_handler(state=None)
async def catchall_no_state(message: types.Message, state: FSMContext):
    """Show welcome screen whenever user sends anything without an active flow."""
    try:
        await message.answer("👇", reply_markup=PERSISTENT_KB)
        await show_welcome(message)
        await MarketFlow.language.set()
    except Exception as exc:
        logger.error("catchall_no_state: %s", exc)

@dp.message_handler(lambda m: m.text and m.text.startswith("/"))
async def unknown_command(message: types.Message, state: FSMContext):
    try:
        data = await state.get_data()
        lang = data.get("language", "en")
        if lang == "ar":
            await message.answer("❓ أمر غير معروف.\nاكتب /start للبدء.")
        else:
            await message.answer("❓ Unknown command.\nType /start to begin.")
    except Exception as exc:
        logger.error("unknown_command: %s", exc)

# ══════════════════════════════════════════════
#  GRACEFUL SHUTDOWN
# ══════════════════════════════════════════════
def _handle_signal(signum, frame):
    logger.info("Received signal %s — shutting down gracefully.", signum)
    sys.exit(0)

signal.signal(signal.SIGTERM, _handle_signal)
signal.signal(signal.SIGINT,  _handle_signal)

# ══════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════
#  ██████╗██████╗ ██╗   ██╗██████╗ ████████╗ ██████╗
# ██╔════╝██╔══██╗╚██╗ ██╔╝██╔══██╗╚══██╔══╝██╔═══██╗
# ██║     ██████╔╝ ╚████╔╝ ██████╔╝   ██║   ██║   ██║
# ██║     ██╔══██╗  ╚██╔╝  ██╔═══╝    ██║   ██║   ██║
# ╚██████╗██║  ██║   ██║   ██║        ██║   ╚██████╔╝
#  ╚═════╝╚═╝  ╚═╝   ╚═╝   ╚═╝        ╚═╝    ╚═════╝
#  TRACKER  MODULE  (inlined from crypto_tracker.py)
# ══════════════════════════════════════════════════════════════════

# ── Tracker DB helpers ────────────────────────────────────────────

def _trk_get_enabled_tokens() -> list:
    conn = sqlite3.connect(DATABASE_PATH)
    c = conn.cursor()
    c.execute("SELECT id, symbol FROM tracked_tokens WHERE enabled=1 ORDER BY symbol")
    rows = c.fetchall()
    conn.close()
    return [{"id": r[0], "symbol": r[1]} for r in rows]


def _trk_get_all_tokens() -> list:
    conn = sqlite3.connect(DATABASE_PATH)
    c = conn.cursor()
    c.execute("SELECT id, symbol, enabled FROM tracked_tokens ORDER BY symbol")
    rows = c.fetchall()
    conn.close()
    return [{"id": r[0], "symbol": r[1], "enabled": bool(r[2])} for r in rows]


def _trk_toggle_token(token_id: str) -> bool:
    conn = sqlite3.connect(DATABASE_PATH)
    c = conn.cursor()
    c.execute("UPDATE tracked_tokens SET enabled = 1 - enabled WHERE id=?", (token_id,))
    conn.commit()
    c.execute("SELECT enabled FROM tracked_tokens WHERE id=?", (token_id,))
    row = c.fetchone()
    conn.close()
    return bool(row[0]) if row else False


def _trk_add_token(token_id: str, symbol: str):
    conn = sqlite3.connect(DATABASE_PATH)
    c = conn.cursor()
    c.execute(
        "INSERT OR IGNORE INTO tracked_tokens (id, symbol) VALUES (?,?)",
        (token_id, symbol.upper())
    )
    conn.commit()
    conn.close()


def _trk_remove_token(token_id: str):
    conn = sqlite3.connect(DATABASE_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM tracked_tokens WHERE id=?", (token_id,))
    conn.commit()
    conn.close()


def _trk_get_user_portfolio(user_id: int) -> list:
    conn = sqlite3.connect(DATABASE_PATH)
    c = conn.cursor()
    c.execute("""
        SELECT p.token_id, t.symbol, p.quantity
        FROM user_portfolio p
        JOIN tracked_tokens t ON p.token_id = t.id
        WHERE p.user_id=? AND p.quantity > 0
        ORDER BY t.symbol
    """, (user_id,))
    rows = c.fetchall()
    conn.close()
    return [{"token_id": r[0], "symbol": r[1], "quantity": r[2]} for r in rows]


def _trk_set_portfolio_qty(user_id: int, token_id: str, quantity: float):
    conn = sqlite3.connect(DATABASE_PATH)
    c = conn.cursor()
    if quantity <= 0:
        c.execute(
            "DELETE FROM user_portfolio WHERE user_id=? AND token_id=?",
            (user_id, token_id)
        )
    else:
        c.execute("""
            INSERT INTO user_portfolio (user_id, token_id, quantity, updated_at)
            VALUES (?,?,?,CURRENT_TIMESTAMP)
            ON CONFLICT(user_id, token_id) DO UPDATE
            SET quantity=excluded.quantity, updated_at=excluded.updated_at
        """, (user_id, token_id, quantity))
    conn.commit()
    conn.close()


# ── CoinGecko API helpers ─────────────────────────────────────────

def _trk_fetch_token_info(token_id: str) -> Optional[dict]:
    """Fetch current price and basic % changes from CoinGecko."""
    try:
        url = f"{COINGECKO_BASE}/coins/{token_id}"
        params = {
            "localization": "false", "tickers": "false",
            "community_data": "false", "developer_data": "false"
        }
        r = requests.get(url, params=params, timeout=10)
        r.raise_for_status()
        data  = r.json()
        mdata = data.get("market_data", {})
        return {
            "name":    data.get("name", token_id),
            "symbol":  data.get("symbol", "").upper(),
            "current_price":          mdata.get("current_price", {}).get("usd", 0) or 0,
            "price_change_24h_pct":   mdata.get("price_change_percentage_24h", 0) or 0,
            "price_change_30d_pct":   mdata.get("price_change_percentage_30d_in_currency", {}).get("usd", 0) or 0,
            "price_change_1y_pct":    mdata.get("price_change_percentage_1y_in_currency",  {}).get("usd", 0) or 0,
        }
    except Exception as e:
        logger.error("CoinGecko info error [%s]: %s", token_id, e)
        return None


def _trk_fetch_historical(token_id: str, days: int) -> list:
    """Fetch daily closing prices for the last N days."""
    try:
        url    = f"{COINGECKO_BASE}/coins/{token_id}/market_chart"
        params = {"vs_currency": "usd", "days": days, "interval": "daily"}
        r = requests.get(url, params=params, timeout=15)
        r.raise_for_status()
        return [p[1] for p in r.json().get("prices", [])]
    except Exception as e:
        logger.error("CoinGecko history error [%s]: %s", token_id, e)
        return []


def _trk_price_at(prices: list, days_ago: int) -> Optional[float]:
    """Return the price N days ago from a daily price list."""
    if not prices or len(prices) < 2:
        return None
    idx = max(0, len(prices) - days_ago - 1)
    return prices[idx]


# ── Technical Analysis ────────────────────────────────────────────

def _trk_calc_rsi(prices: list, period: int = 14) -> Optional[float]:
    if len(prices) < period + 1:
        return None
    gains, losses = [], []
    for i in range(1, period + 1):
        diff = prices[i] - prices[i - 1]
        gains.append(max(diff, 0))
        losses.append(max(-diff, 0))
    avg_gain = sum(gains) / period
    avg_loss = sum(losses) / period
    if avg_loss == 0:
        return 100.0
    rs = avg_gain / avg_loss
    return 100 - (100 / (1 + rs))


def _trk_calc_sma(prices: list, period: int) -> Optional[float]:
    if len(prices) < period:
        return None
    return sum(prices[-period:]) / period


def _trk_build_analysis(token_id: str) -> dict:
    """Fetch 730 days of data and compute RSI, MA50/200, momentum."""
    prices_730 = _trk_fetch_historical(token_id, 730)
    prices_200 = prices_730[-200:] if len(prices_730) >= 200 else prices_730

    result = {
        "prices_730": prices_730,
        "rsi":        None,
        "ma50":       None,
        "ma200":      None,
        "ma_signal":  None,
        "momentum_90d": None,
    }
    if prices_200:
        rsi_window = prices_200[-15:] if len(prices_200) >= 15 else prices_200
        result["rsi"]   = _trk_calc_rsi(rsi_window)
        result["ma50"]  = _trk_calc_sma(prices_200, 50)
        result["ma200"] = _trk_calc_sma(prices_200, 200)
        if result["ma50"] and result["ma200"]:
            result["ma_signal"] = "bullish" if result["ma50"] > result["ma200"] else "bearish"

    if prices_730 and len(prices_730) >= 90:
        p_now = prices_730[-1]
        p_90  = prices_730[-90]
        result["momentum_90d"] = ((p_now - p_90) / p_90 * 100) if p_90 else None

    return result


def _trk_rule_verdict(rsi, ma_signal, momentum_90d, pct_1d, pct_30d):
    score   = 0
    reasons = []

    if rsi is not None:
        if rsi < 30:
            score += 2; reasons.append("RSI oversold — strong buy signal")
        elif rsi < 45:
            score += 1; reasons.append("RSI below midpoint — mild bullish")
        elif rsi > 70:
            score -= 2; reasons.append("RSI overbought — caution / sell signal")
        elif rsi > 55:
            score -= 1; reasons.append("RSI above midpoint — mild bearish")

    if ma_signal == "bullish":
        score += 2; reasons.append("MA50 > MA200 (Golden Cross — bullish)")
    elif ma_signal == "bearish":
        score -= 2; reasons.append("MA50 < MA200 (Death Cross — bearish)")

    if momentum_90d is not None:
        if momentum_90d > 20:
            score += 1; reasons.append(f"Strong 90D momentum (+{momentum_90d:.1f}%)")
        elif momentum_90d < -20:
            score -= 1; reasons.append(f"Weak 90D momentum ({momentum_90d:.1f}%)")

    if pct_30d is not None:
        if pct_30d > 10:  score += 1
        elif pct_30d < -10: score -= 1

    if score >= 3:    verdict = "🟢 STRONG BUY"
    elif score >= 1:  verdict = "🟢 BUY / ACCUMULATE"
    elif score == 0:  verdict = "🟡 HOLD / NEUTRAL"
    elif score >= -2: verdict = "🔴 CAUTION / REDUCE"
    else:             verdict = "🔴 STRONG SELL"

    explanation = " • ".join(reasons[:3]) if reasons else "Insufficient data for full analysis."
    return verdict, explanation


async def _trk_openai_analysis(
    token_name: str, symbol: str, current_price: float,
    pct_1d: float, pct_30d: float, pct_90d: float,
    pct_6m: float, pct_1y: float, pct_2y: float,
    rsi: Optional[float], ma_signal: Optional[str],
    lang: str = "en"
) -> str:
    if not OPENAI_API_KEY:
        return "⚠️ OpenAI API key not configured. Add OPENAI_API_KEY to your .env file."
    try:
        import openai
        client = openai.AsyncOpenAI(api_key=OPENAI_API_KEY)
        lang_instruction = (
            "Respond in Arabic (العربية)." if lang == "ar"
            else "Respond in English."
        )
        prompt = (
            f"You are a professional crypto market analyst. {lang_instruction} "
            f"Analyze {token_name} ({symbol}) and give a clear buy/sell/hold recommendation.\n\n"
            f"Current Data:\n"
            f"- Price: ${current_price:,.4f}\n"
            f"- 1D:  {pct_1d:+.2f}%\n"
            f"- 30D: {pct_30d:+.2f}%\n"
            f"- 90D: {pct_90d:+.2f}%\n"
            f"- 6M:  {pct_6m:+.2f}%\n"
            f"- 1Y:  {pct_1y:+.2f}%\n"
            f"- 2Y:  {pct_2y:+.2f}%\n"
            f"- RSI(14): {f'{rsi:.1f}' if rsi else 'N/A'}\n"
            f"- MA Signal: {ma_signal or 'N/A'}\n\n"
            f"Write a concise 3-4 sentence analysis with a clear BUY / HOLD / SELL verdict. "
            f"Focus on key patterns, momentum, and risk. Keep it simple and actionable."
        )
        response = await client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=220,
            temperature=0.7,
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        logger.error("OpenAI error: %s", e)
        return f"AI analysis unavailable: {e}"


# ── Formatters ────────────────────────────────────────────────────

def _trk_fmt_price(price: float) -> str:
    if price >= 1000:    return f"${price:,.2f}"
    elif price >= 1:     return f"${price:.4f}"
    elif price >= 0.01:  return f"${price:.5f}"
    else:                return f"${price:.8f}"


def _trk_fmt_pct(pct: Optional[float]) -> str:
    if pct is None: return "  N/A      "
    arrow = "🟢" if pct >= 0 else "🔴"
    sign  = "+" if pct >= 0 else ""
    return f"{arrow} {sign}{pct:.2f}%"


def _trk_fmt_delta(now: float, old: Optional[float]) -> str:
    if old is None or old == 0: return "N/A"
    diff = now - old
    pct  = (diff / old) * 100
    sign = "+" if diff >= 0 else ""
    arrow = "🟢" if diff >= 0 else "🔴"
    return f"{arrow} {sign}{_trk_fmt_price(abs(diff))}  ({sign}{pct:.2f}%)"


def _trk_fmt_rsi(rsi: Optional[float]) -> str:
    if rsi is None: return "N/A"
    if rsi < 30:    label = "Oversold  🟢"
    elif rsi > 70:  label = "Overbought 🔴"
    else:           label = "Neutral   🟡"
    return f"{rsi:.1f} — {label}"


# ── Text Builders ─────────────────────────────────────────────────

def _trk_build_token_text(token_id: str, info: dict, prices_730: list) -> str:
    emoji   = TOKEN_EMOJIS.get(token_id, "🔵")
    name    = info["name"]
    symbol  = info["symbol"]
    current = info["current_price"]

    def row(label, days_ago, override_pct=None):
        old = _trk_price_at(prices_730, days_ago)
        old_str = _trk_fmt_price(old) if old else "N/A        "
        pct = override_pct if override_pct is not None else (
            ((current - old) / old * 100) if old and old > 0 else None
        )
        return f" {label:<11} {old_str:<15} {_trk_fmt_pct(pct)}"

    now_str = datetime.now().strftime("%d %b %Y, %H:%M")

    return (
        f"╔══════════════════════════════════╗\n"
        f"║  {emoji}  {name} ({symbol})\n"
        f"╚══════════════════════════════════╝\n\n"
        f"💵 Current:  {_trk_fmt_price(current)}\n\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f" ⏱ Period     📌 Was           📉📈 Change\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"{row('1 Day',     1,   info.get('price_change_24h_pct'))}\n"
        f"{row('1 Month',   30,  info.get('price_change_30d_pct'))}\n"
        f"{row('3 Months',  90)}\n"
        f"{row('6 Months',  180)}\n"
        f"{row('1 Year',    365, info.get('price_change_1y_pct'))}\n"
        f"{row('2 Years',   730)}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🕐 Updated: {now_str}"
    )


def _trk_build_analysis_text(symbol, name, rsi, ma50, ma200,
                               ma_signal, momentum_90d, pct_1d, pct_30d,
                               verdict, explanation, lang: str = "en") -> str:
    ma50_str  = _trk_fmt_price(ma50)  if ma50  else "N/A"
    ma200_str = _trk_fmt_price(ma200) if ma200 else "N/A"

    if lang == "ar":
        ma_str = ("🟢 تقاطع صاعد" if ma_signal == "bullish"
                  else "🔴 تقاطع هابط" if ma_signal == "bearish" else "غير متاح")
        return (
            f"╔══════════════════════════════════╗\n"
            f"║  🤖 تحليل ذكي — {symbol} ({name})\n"
            f"╚══════════════════════════════════╝\n\n"
            f"📊 المؤشرات الفنية:\n"
            f"  RSI (14):       {_trk_fmt_rsi(rsi)}\n"
            f"  MA 50:          {ma50_str}\n"
            f"  MA 200:         {ma200_str}\n"
            f"  إشارة MA:       {ma_str}\n"
            f"  زخم 90 يوم:     {_trk_fmt_pct(momentum_90d)}\n"
            f"  تغير 1 يوم:     {_trk_fmt_pct(pct_1d)}\n"
            f"  تغير 30 يوم:    {_trk_fmt_pct(pct_30d)}\n\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"🤖 حكم النظام:\n\n"
            f"  {verdict}\n\n"
            f'  \"{explanation}\"\n\n'
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"⚠️ هذا ليس نصيحة مالية."
        )
    else:
        ma_str = ("🟢 Bullish Cross" if ma_signal == "bullish"
                  else "🔴 Bearish Cross" if ma_signal == "bearish" else "N/A")
        return (
            f"╔══════════════════════════════════╗\n"
            f"║  🤖 AI Analysis — {symbol} ({name})\n"
            f"╚══════════════════════════════════╝\n\n"
            f"📊 Technical Signals:\n"
            f"  RSI (14):       {_trk_fmt_rsi(rsi)}\n"
            f"  MA 50:          {ma50_str}\n"
            f"  MA 200:         {ma200_str}\n"
            f"  MA Signal:      {ma_str}\n"
            f"  Momentum 90D:   {_trk_fmt_pct(momentum_90d)}\n"
            f"  1D Change:      {_trk_fmt_pct(pct_1d)}\n"
            f"  30D Change:     {_trk_fmt_pct(pct_30d)}\n\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"🤖 Rule-Based Verdict:\n\n"
            f"  {verdict}\n\n"
            f'  \"{explanation}\"\n\n'
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"⚠️ Not financial advice."
        )


def _trk_build_portfolio_text(user_id: int, holdings: list,
                               prices_map: dict, hist_map: dict,
                               lang: str = "en") -> str:
    lines    = []
    total    = 0.0
    if lang == "ar":
        periods = {"يوم 1": 1, "شهر 1": 30, "6 أشهر": 180, "سنة 1": 365, "سنتان": 730}
    else:
        periods = {"1 Day": 1, "1 Month": 30, "6 Months": 180, "1 Year": 365, "2 Years": 730}
    hist_totals = {k: 0.0 for k in periods}
    hist_valid  = {k: True  for k in periods}

    for h in holdings:
        info  = prices_map.get(h["token_id"])
        if not info:
            continue
        price  = info["current_price"]
        value  = price * h["quantity"]
        total += value
        emoji  = TOKEN_EMOJIS.get(h["token_id"], "🔵")
        qty_str = f"{h['quantity']:g}"
        lines.append(
            f" {emoji} {h['symbol']:<6}  {qty_str:<10} {_trk_fmt_price(price):<16} {_trk_fmt_price(value)}"
        )
        hist_prices = hist_map.get(h["token_id"], [])
        for label, days in periods.items():
            old_price = _trk_price_at(hist_prices, days)
            if old_price and old_price > 0:
                hist_totals[label] += old_price * h["quantity"]
            else:
                hist_valid[label] = False

    now_str    = datetime.now().strftime("%d %b %Y, %H:%M")
    holdings_block = "\n".join(lines) if lines else " (no holdings)"

    perf_rows = []
    for label, days in periods.items():
        old_val = hist_totals[label] if hist_valid[label] else None
        perf_rows.append(f" {label:<11}  {_trk_fmt_delta(total, old_val)}")

    perf_block = "\n".join(perf_rows)

    if lang == "ar":
        return (
            f"╔══════════════════════════════════╗\n"
            f"║   💼  محفظتي                     ║\n"
            f"╚══════════════════════════════════╝\n\n"
            f" العملة   الكمية     السعر            القيمة\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"{holdings_block}\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f" 💰 الإجمالي الآن:  {_trk_fmt_price(total)}\n\n"
            f"📈 أداء المحفظة:\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f" ⏱ الفترة      📉📈 التغيير (USD + %)\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"{perf_block}\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"🕐 {now_str}\n\n"
            f"اضغط على عملة لتعديل الكمية:"
        )
    else:
        return (
            f"╔══════════════════════════════════╗\n"
            f"║   💼  MY PORTFOLIO               ║\n"
            f"╚══════════════════════════════════╝\n\n"
            f" Token    Qty        Price            Value\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"{holdings_block}\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f" 💰 Total Now:   {_trk_fmt_price(total)}\n\n"
            f"📈 Portfolio Performance:\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f" ⏱ Period      📉📈 Change (USD + %)\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"{perf_block}\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"🕐 {now_str}\n\n"
            f"Tap a token below to edit quantity:"
        )


# ── Tracker Keyboards ─────────────────────────────────────────────

def _trk_tokens_list_kb(tokens: list, is_admin: bool) -> InlineKeyboardMarkup:
    buttons = []
    row = []
    for tok in tokens:
        emoji = TOKEN_EMOJIS.get(tok["id"], "🔵")
        row.append(InlineKeyboardButton(
            text=f"{emoji} {tok['symbol']}",
            callback_data=f"trk:token:{tok['id']}"
        ))
        if len(row) == 3:
            buttons.append(row); row = []
    if row:
        buttons.append(row)
    buttons.append([InlineKeyboardButton(text="💼 My Portfolio", callback_data="trk:portfolio")])
    buttons.append([InlineKeyboardButton(text="📥 Export to Excel", callback_data="trk:export")])
    if is_admin:
        buttons.append([InlineKeyboardButton(text="⚙️ Manage Tokens", callback_data="trk:admin")])
    buttons.append([InlineKeyboardButton(text="❌ Close", callback_data="trk:close")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def _trk_token_detail_kb(token_id: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="🤖 AI Analysis",  callback_data=f"trk:ai:{token_id}"),
            InlineKeyboardButton(text="🔄 Refresh",       callback_data=f"trk:token:{token_id}"),
        ],
        [
            InlineKeyboardButton(text="💼 My Portfolio", callback_data="trk:portfolio"),
            InlineKeyboardButton(text="◀️ Back",          callback_data="trk:list"),
        ],
    ])


def _trk_analysis_kb(token_id: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🧠 Deep AI Analysis (GPT-4)", callback_data=f"trk:gpt:{token_id}")],
        [
            InlineKeyboardButton(text="🔄 Refresh",  callback_data=f"trk:ai:{token_id}"),
            InlineKeyboardButton(text="◀️ Back",      callback_data=f"trk:token:{token_id}"),
        ],
    ])


def _trk_portfolio_kb(tokens: list) -> InlineKeyboardMarkup:
    buttons = []
    row = []
    for tok in tokens:
        emoji = TOKEN_EMOJIS.get(tok["id"], "🔵")
        row.append(InlineKeyboardButton(
            text=f"{emoji} {tok['symbol']}",
            callback_data=f"trk:pf_edit:{tok['id']}"
        ))
        if len(row) == 3:
            buttons.append(row); row = []
    if row:
        buttons.append(row)
    buttons.append([
        InlineKeyboardButton(text="🔄 Refresh",  callback_data="trk:portfolio"),
        InlineKeyboardButton(text="◀️ Back",      callback_data="trk:list"),
    ])
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def _trk_admin_tokens_kb(tokens: list) -> InlineKeyboardMarkup:
    buttons = []
    for tok in tokens:
        status = "✅" if tok["enabled"] else "❌"
        buttons.append([InlineKeyboardButton(
            text=f"{status} {tok['symbol']}  ({tok['id']})",
            callback_data=f"trk:admin_toggle:{tok['id']}"
        )])
    buttons.append([
        InlineKeyboardButton(text="📋 Popular Tokens", callback_data="trk:admin_add_popular"),
        InlineKeyboardButton(text="🔍 Search Token",   callback_data="trk:admin_search"),
    ])
    buttons.append([InlineKeyboardButton(text="◀️ Back",      callback_data="trk:list")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def _trk_popular_tokens_kb(existing_ids: set) -> InlineKeyboardMarkup:
    buttons = []
    row = []
    for tid, sym in POPULAR_TOKENS:
        if tid not in existing_ids:
            row.append(InlineKeyboardButton(
                text=sym,
                callback_data=f"trk:admin_add:{tid}:{sym}"
            ))
            if len(row) == 4:
                buttons.append(row); row = []
    if row:
        buttons.append(row)
    buttons.append([InlineKeyboardButton(text="◀️ Back", callback_data="trk:admin")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def _trk_cancel_kb(back_cb: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="❌ Cancel", callback_data=back_cb)]
    ])


def _trk_search_results_kb(results: list, existing_ids: set) -> InlineKeyboardMarkup:
    """Keyboard showing CoinGecko search results for admin to pick from."""
    buttons = []
    for coin in results[:10]:
        cid = coin.get("id", "")
        sym = coin.get("symbol", "").upper()
        name = coin.get("name", "")
        if cid in existing_ids:
            label = f"✅ {sym} – {name}"
            cb_data = f"trk:search_already:{cid}"
        else:
            label = f"➕ {sym} – {name}"
            cb_data = f"trk:admin_add:{cid}:{sym}"
        buttons.append([InlineKeyboardButton(text=label[:60], callback_data=cb_data)])
    buttons.append([InlineKeyboardButton(text="🔍 Search Again", callback_data="trk:admin_search")])
    buttons.append([InlineKeyboardButton(text="◀️ Back",        callback_data="trk:admin")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)


async def _trk_coingecko_search(query: str) -> list:
    """Search CoinGecko for coins matching query. Returns list of {id, symbol, name}."""
    try:
        url = f"{COINGECKO_BASE}/search?query={query}"
        async with aiohttp.ClientSession() as session:
            async with session.get(url, timeout=aiohttp.ClientTimeout(total=10)) as resp:
                if resp.status != 200:
                    return []
                data = await resp.json()
                return data.get("coins", [])
    except Exception as exc:
        logger.error("CoinGecko search error: %s", exc)
        return []


# ── Excel Export ──────────────────────────────────────────────────

def _trk_build_excel(tokens_data: list) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Crypto Prices"

    DARK_BG    = "1A1A2E"; MID_BG   = "16213E"; ROW_EVEN  = "0F3460"; ROW_ODD   = "162447"
    GREEN_FILL = "1B5E20"; RED_FILL = "7F0000"; GREEN_TXT = "A5D6A7"; RED_TXT   = "EF9A9A"
    WHITE = "FFFFFF"; GOLD = "FFD700"; SILVER = "C0C0C0"
    thin   = Side(style="thin", color="2A2A4A")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell_style(cell, bold=False, font_color=WHITE, bg=None, align="center", size=10):
        cell.font      = Font(bold=bold, color=font_color, size=size, name="Calibri")
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        cell.border    = border
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)

    def pct_style(cell, pct):
        if pct is None:
            cell.value = "N/A"
            cell_style(cell, bg=ROW_EVEN if cell.row % 2 == 0 else ROW_ODD, font_color=SILVER)
            return
        sign = "+" if pct >= 0 else ""
        cell.value = f"{sign}{pct:.2f}%"
        if pct >= 0:
            cell_style(cell, bold=True, bg=GREEN_FILL, font_color=GREEN_TXT)
        else:
            cell_style(cell, bold=True, bg=RED_FILL, font_color=RED_TXT)

    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = f"📊 Crypto Price Report  —  {datetime.now().strftime('%d %b %Y  %H:%M')} UTC"
    cell_style(title_cell, bold=True, bg=DARK_BG, font_color=GOLD, align="center", size=13)
    ws.row_dimensions[1].height = 28

    headers = ["#", "Token", "Symbol", "Price (USD)", "1 Day", "1 Month", "3 Months", "6 Months", "1 Year", "2 Years"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=h)
        cell_style(c, bold=True, bg=MID_BG, font_color=GOLD, size=10)
    ws.row_dimensions[2].height = 22

    col_widths = [5, 18, 9, 16, 12, 12, 12, 12, 12, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for idx, td in enumerate(tokens_data, start=1):
        row_n = idx + 2
        bg    = ROW_EVEN if idx % 2 == 0 else ROW_ODD
        emoji = TOKEN_EMOJIS.get(td["token_id"], "")

        c = ws.cell(row=row_n, column=1, value=idx)
        cell_style(c, bg=bg, font_color=SILVER)
        c = ws.cell(row=row_n, column=2, value=f"{emoji}  {td['name']}")
        cell_style(c, bold=True, bg=bg, font_color=WHITE, align="left")
        c = ws.cell(row=row_n, column=3, value=td["symbol"])
        cell_style(c, bold=True, bg=bg, font_color=GOLD)

        price = td["current_price"]
        if price >= 1000:   price_str = f"${price:,.2f}"
        elif price >= 1:    price_str = f"${price:.4f}"
        elif price >= 0.01: price_str = f"${price:.5f}"
        else:               price_str = f"${price:.8f}"
        c = ws.cell(row=row_n, column=4, value=price_str)
        cell_style(c, bold=True, bg=bg, font_color=WHITE)

        for col_n, key in enumerate(["pct_1d", "pct_1m", "pct_3m", "pct_6m", "pct_1y", "pct_2y"], start=5):
            pct_style(ws.cell(row=row_n, column=col_n), td.get(key))

        ws.row_dimensions[row_n].height = 20

    footer_row = len(tokens_data) + 3
    ws.merge_cells(f"A{footer_row}:J{footer_row}")
    fc = ws.cell(row=footer_row, column=1, value="⚠️  Data sourced from CoinGecko. Not financial advice.")
    cell_style(fc, bg=DARK_BG, font_color=SILVER, align="center", size=9)
    ws.row_dimensions[footer_row].height = 18
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _trk_fetch_all_for_export(tokens: list) -> list:
    """Blocking: fetch all token data for Excel export."""
    result = []
    for tok in tokens:
        info   = _trk_fetch_token_info(tok["id"])
        prices = _trk_fetch_historical(tok["id"], 730)
        if not info:
            continue
        current = info["current_price"]

        def spct(days, override=None):
            if override is not None:
                return override
            old = _trk_price_at(prices, days)
            return ((current - old) / old * 100) if old and old > 0 else None

        result.append({
            "token_id":      t["id"],
            "name":          info["name"],
            "symbol":        info["symbol"],
            "current_price": current,
            "pct_1d":        spct(1,   info.get("price_change_24h_pct")),
            "pct_1m":        spct(30,  info.get("price_change_30d_pct")),
            "pct_3m":        spct(90),
            "pct_6m":        spct(180),
            "pct_1y":        spct(365, info.get("price_change_1y_pct")),
            "pct_2y":        spct(730),
        })
    return result


# ── Tracker Handlers (aiogram 2.x) ───────────────────────────────

async def cmd_tracker(message: types.Message):
    """Show crypto tracker menu."""
    try:
        if not await asyncio.to_thread(is_section_enabled, "section_crypto"):
            await message.answer(SECTION_CLOSED_MSG["section_crypto"])
            return
        if is_rate_limited(message.from_user.id):
            return
        tokens   = await asyncio.to_thread(_trk_get_enabled_tokens)
        is_admin = message.from_user.id == ADMIN_ID
        if not tokens:
            if is_admin:
                msg = "📊 *Crypto Tracker*\n\n⚠️ No tokens added yet.\nTap ⚙️ Manage Tokens to add some."
            else:
                msg = "📊 *Crypto Tracker*\n\n⚠️ No tokens available yet.\nPlease check back later."
            await message.answer(
                msg,
                parse_mode="Markdown",
                reply_markup=_trk_tokens_list_kb([], is_admin)
            )
            return
        await message.answer(
            "📊 *Crypto Tracker*\n\nSelect a token to view details:",
            parse_mode="Markdown",
            reply_markup=_trk_tokens_list_kb(tokens, is_admin)
        )
    except Exception as exc:
        logger.error("cmd_tracker: %s", exc)
        try:
            await message.answer(f"❌ Error loading Crypto Tracker: {exc}")
        except Exception:
            pass


async def cb_trk_list(cb: types.CallbackQuery):
    tokens   = await asyncio.to_thread(_trk_get_enabled_tokens)
    is_admin = cb.from_user.id == ADMIN_ID
    await cb.message.edit_text(
        "📊 *Crypto Tracker*\n\nSelect a token to view details:",
        parse_mode="Markdown",
        reply_markup=_trk_tokens_list_kb(tokens, is_admin)
    )
    await cb.answer()


async def cb_trk_token(cb: types.CallbackQuery):
    token_id = cb.data.split(":", 2)[2]
    await cb.answer("⏳ Fetching prices...")
    await cb.message.edit_text("⏳ Loading price data, please wait...")

    info, prices = await asyncio.gather(
        asyncio.to_thread(_trk_fetch_token_info, token_id),
        asyncio.to_thread(_trk_fetch_historical, token_id, 730),
    )
    if not info:
        await cb.message.edit_text(
            "❌ Could not fetch data. Please try again.",
            reply_markup=_trk_token_detail_kb(token_id)
        )
        return

    text = _trk_build_token_text(token_id, info, prices)
    await cb.message.edit_text(text, reply_markup=_trk_token_detail_kb(token_id))


async def cb_trk_ai(cb: types.CallbackQuery):
    token_id = cb.data.split(":", 2)[2]
    await cb.answer("🤖 Analyzing...")
    await cb.message.edit_text("🤖 Running technical analysis, please wait...")

    info, analysis = await asyncio.gather(
        asyncio.to_thread(_trk_fetch_token_info, token_id),
        asyncio.to_thread(_trk_build_analysis, token_id),
    )
    if not info:
        await cb.message.edit_text("❌ Failed to fetch data.", reply_markup=_trk_analysis_kb(token_id))
        return

    user_lang = await asyncio.to_thread(get_user_lang, cb.from_user.id)
    verdict, explanation = _trk_rule_verdict(
        analysis["rsi"], analysis["ma_signal"], analysis["momentum_90d"],
        info.get("price_change_24h_pct"), info.get("price_change_30d_pct")
    )
    text = _trk_build_analysis_text(
        info["symbol"], info["name"],
        analysis["rsi"], analysis["ma50"], analysis["ma200"],
        analysis["ma_signal"], analysis["momentum_90d"],
        info.get("price_change_24h_pct"), info.get("price_change_30d_pct"),
        verdict, explanation, lang=user_lang
    )
    await cb.message.edit_text(text, reply_markup=_trk_analysis_kb(token_id))


async def cb_trk_gpt(cb: types.CallbackQuery):
    token_id = cb.data.split(":", 2)[2]
    if not OPENAI_API_KEY:
        await cb.answer("⚠️ OpenAI API key not configured!", show_alert=True)
        return
    await cb.answer("🧠 Asking GPT-4...")
    await cb.message.edit_text("🧠 GPT-4 is analyzing market data, please wait...")

    info, analysis = await asyncio.gather(
        asyncio.to_thread(_trk_fetch_token_info, token_id),
        asyncio.to_thread(_trk_build_analysis, token_id),
    )
    if not info:
        await cb.message.edit_text("❌ Failed to fetch data.", reply_markup=_trk_analysis_kb(token_id))
        return

    prices_730 = analysis.get("prices_730", [])
    current    = info["current_price"]

    def spct(days):
        old = _trk_price_at(prices_730, days)
        return ((current - old) / old * 100) if old and old > 0 else 0.0

    user_lang = await asyncio.to_thread(get_user_lang, cb.from_user.id)
    gpt_text = await _trk_openai_analysis(
        info["name"], info["symbol"], current,
        info.get("price_change_24h_pct", 0),
        info.get("price_change_30d_pct", 0),
        spct(90), spct(180), spct(365), spct(730),
        analysis["rsi"], analysis["ma_signal"],
        lang=user_lang
    )

    disclaimer = "⚠️ هذا ليس نصيحة مالية." if user_lang == "ar" else "⚠️ Not financial advice."
    header     = "🧠 تحليل GPT-4" if user_lang == "ar" else "🧠 GPT-4 Analysis"
    text = (
        f"╔══════════════════════════════════╗\n"
        f"║  {header} — {info['symbol']}\n"
        f"╚══════════════════════════════════╝\n\n"
        f"{gpt_text}\n\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"{disclaimer}"
    )
    await cb.message.edit_text(text, reply_markup=_trk_analysis_kb(token_id))


async def cb_trk_portfolio(cb: types.CallbackQuery):
    user_id = cb.from_user.id
    await cb.answer("⏳ Loading portfolio...")
    await cb.message.edit_text("⏳ Loading your portfolio, please wait...")

    holdings       = await asyncio.to_thread(_trk_get_user_portfolio, user_id)
    enabled_tokens = await asyncio.to_thread(_trk_get_enabled_tokens)

    if not holdings:
        await cb.message.edit_text(
            "💼 *My Portfolio*\n\nYou have no holdings yet.\n"
            "Tap a token below to set your quantity:",
            parse_mode="Markdown",
            reply_markup=_trk_portfolio_kb(enabled_tokens)
        )
        return

    token_ids     = [h["token_id"] for h in holdings]
    price_tasks   = [asyncio.to_thread(_trk_fetch_token_info, tid)          for tid in token_ids]
    history_tasks = [asyncio.to_thread(_trk_fetch_historical, tid, 730)     for tid in token_ids]

    price_results, history_results = await asyncio.gather(
        asyncio.gather(*price_tasks),
        asyncio.gather(*history_tasks),
    )

    prices_map = {tid: info for tid, info in zip(token_ids, price_results) if info}
    hist_map   = {tid: hist for tid, hist in zip(token_ids, history_results)}

    user_lang = await asyncio.to_thread(get_user_lang, user_id)
    text = _trk_build_portfolio_text(user_id, holdings, prices_map, hist_map, lang=user_lang)
    await cb.message.edit_text(text, reply_markup=_trk_portfolio_kb(enabled_tokens))


async def cb_pf_edit(cb: types.CallbackQuery, state: FSMContext):
    user_id  = cb.from_user.id
    token_id = cb.data.split(":", 2)[2]
    tokens   = await asyncio.to_thread(_trk_get_enabled_tokens)
    token    = next((t for t in tokens if t["id"] == token_id), None)
    if not token:
        await cb.answer("Token not found.", show_alert=True)
        return

    await PortfolioStates.entering_quantity.set()
    await state.update_data(token_id=token_id, symbol=token["symbol"])

    emoji = TOKEN_EMOJIS.get(token_id, "🔵")
    await cb.message.edit_text(
        f"✏️ Enter quantity for {emoji} *{token['symbol']}*\n\n"
        f"Type the amount you hold (e.g. `0.5` or `1500`).\n"
        f"Send `0` to remove from portfolio.",
        parse_mode="Markdown",
        reply_markup=_trk_cancel_kb("trk:portfolio")
    )
    await cb.answer()


async def handle_portfolio_qty(message: types.Message, state: FSMContext):
    data     = await state.get_data()
    token_id = data.get("token_id", "")
    symbol   = data.get("symbol", "")

    try:
        qty = float(message.text.strip().replace(",", "."))
        if qty < 0:
            raise ValueError("negative")
    except (ValueError, AttributeError):
        await message.answer(
            "❌ Invalid amount. Please enter a positive number like `0.5` or `1500`.",
            parse_mode="Markdown"
        )
        return

    await asyncio.to_thread(_trk_set_portfolio_qty, message.from_user.id, token_id, qty)
    await state.finish()

    if qty == 0:
        await message.answer(f"🗑️ Removed *{symbol}* from your portfolio.", parse_mode="Markdown")
    else:
        await message.answer(
            f"✅ Updated *{symbol}* → `{qty:g}` saved in your portfolio.",
            parse_mode="Markdown"
        )


async def cb_trk_admin(cb: types.CallbackQuery):
    if cb.from_user.id != ADMIN_ID:
        await cb.answer("⛔ Admin only.", show_alert=True)
        return
    tokens = await asyncio.to_thread(_trk_get_all_tokens)
    await cb.message.edit_text(
        "⚙️ *Manage Tracked Tokens*\n\n"
        "Tap a token to enable ✅ / disable ❌ it for all users.\n"
        "Add new tokens with the button below.",
        parse_mode="Markdown",
        reply_markup=_trk_admin_tokens_kb(tokens)
    )
    await cb.answer()


async def cb_admin_toggle(cb: types.CallbackQuery):
    if cb.from_user.id != ADMIN_ID:
        await cb.answer("⛔ Admin only.", show_alert=True)
        return
    token_id  = cb.data.split(":", 2)[2]
    new_state = await asyncio.to_thread(_trk_toggle_token, token_id)
    await cb.answer("✅ Enabled" if new_state else "❌ Disabled")
    tokens = await asyncio.to_thread(_trk_get_all_tokens)
    await cb.message.edit_reply_markup(reply_markup=_trk_admin_tokens_kb(tokens))


async def cb_admin_add_popular(cb: types.CallbackQuery):
    if cb.from_user.id != ADMIN_ID:
        await cb.answer("⛔ Admin only.", show_alert=True)
        return
    all_tokens   = await asyncio.to_thread(_trk_get_all_tokens)
    existing_ids = {t["id"] for t in all_tokens}
    await cb.message.edit_text(
        "➕ *Add Token*\n\nSelect from popular tokens:",
        parse_mode="Markdown",
        reply_markup=_trk_popular_tokens_kb(existing_ids)
    )
    await cb.answer()


async def cb_admin_add_token(cb: types.CallbackQuery):
    if cb.from_user.id != ADMIN_ID:
        await cb.answer("⛔ Admin only.", show_alert=True)
        return
    parts    = cb.data.split(":")   # trk:admin_add:token_id:SYMBOL
    token_id = parts[2]
    symbol   = parts[3]
    await asyncio.to_thread(_trk_add_token, token_id, symbol)
    await cb.answer(f"✅ {symbol} added!")
    tokens = await asyncio.to_thread(_trk_get_all_tokens)
    await cb.message.edit_text(
        "⚙️ *Manage Tracked Tokens*\n\n"
        "Tap a token to enable ✅ / disable ❌ it for all users.",
        parse_mode="Markdown",
        reply_markup=_trk_admin_tokens_kb(tokens)
    )


async def cb_admin_search_prompt(cb: types.CallbackQuery, state: FSMContext):
    """Admin taps 🔍 Search Token — ask them to type a coin name or symbol."""
    if cb.from_user.id != ADMIN_ID:
        await cb.answer("⛔ Admin only.", show_alert=True)
        return
    await AdminTokenSearch.waiting_query.set()
    await cb.message.edit_text(
        "🔍 *Search for a Token*\n\n"
        "Type the coin name or ticker symbol:\n"
        "_(e.g.  `pepe`, `sei`, `injective`, `wif`)_",
        parse_mode="Markdown",
        reply_markup=_trk_cancel_kb("trk:admin")
    )
    await cb.answer()


async def on_admin_token_search(message: types.Message, state: FSMContext):
    """Admin typed a search query — search CoinGecko and display results."""
    if message.from_user.id != ADMIN_ID:
        return
    query = message.text.strip()
    if not query:
        await message.answer("❗ Please type a token name or symbol.")
        return
    await state.finish()
    searching_msg = await message.answer(f"🔍 Searching for *{query}*…", parse_mode="Markdown")
    results = await _trk_coingecko_search(query)
    if not results:
        await searching_msg.edit_text(
            f"❌ No results found for *{query}*.\nTry a different name or symbol.",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔍 Search Again", callback_data="trk:admin_search")],
                [InlineKeyboardButton(text="◀️ Back",         callback_data="trk:admin")],
            ])
        )
        return
    all_tokens   = await asyncio.to_thread(_trk_get_all_tokens)
    existing_ids = {t["id"] for t in all_tokens}
    await searching_msg.edit_text(
        f"🔍 Results for *{query}* — tap to add:",
        parse_mode="Markdown",
        reply_markup=_trk_search_results_kb(results, existing_ids)
    )


async def cb_search_already(cb: types.CallbackQuery):
    """Admin tapped an already-added token in search results."""
    await cb.answer("✅ This token is already in your list!", show_alert=True)


async def cb_trk_close(cb: types.CallbackQuery):
    """Delete the crypto tracker message when admin/user taps ❌ Close."""
    try:
        await cb.message.delete()
    except Exception:
        await cb.answer("Closed.")


async def cb_trk_export(cb: types.CallbackQuery):
    await cb.answer("📥 Preparing Excel file...")
    await cb.message.edit_text("⏳ Fetching latest prices for all tokens, please wait...")

    tokens = await asyncio.to_thread(_trk_get_enabled_tokens)
    if not tokens:
        await cb.message.edit_text(
            "❌ No tokens to export. Add tokens first via ⚙️ Manage Tokens.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                InlineKeyboardButton(text="◀️ Back", callback_data="trk:list")
            ]])
        )
        return

    try:
        tokens_data = await asyncio.to_thread(_trk_fetch_all_for_export, tokens)
        xlsx_bytes  = await asyncio.to_thread(_trk_build_excel, tokens_data)
    except Exception as e:
        logger.error("Tracker export error: %s", e)
        await cb.message.edit_text(
            f"❌ Export failed: {e}",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                InlineKeyboardButton(text="◀️ Back", callback_data="trk:list")
            ]])
        )
        return

    filename = f"crypto_prices_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    # aiogram 2.x: use types.InputFile with BytesIO
    file = types.InputFile(io.BytesIO(xlsx_bytes), filename=filename)

    await cb.message.answer_document(
        document=file,
        caption=(
            f"📊 *Crypto Price Report*\n"
            f"📅 {datetime.now().strftime('%d %b %Y, %H:%M')} UTC\n"
            f"📈 {len(tokens_data)} tokens exported\n\n"
            f"_Includes: Price, 1D / 1M / 3M / 6M / 1Y / 2Y changes_"
        ),
        parse_mode="Markdown"
    )

    is_admin = cb.from_user.id == ADMIN_ID
    await cb.message.edit_text(
        "📊 *Crypto Tracker*\n\nSelect a token to view details:",
        parse_mode="Markdown",
        reply_markup=_trk_tokens_list_kb(tokens, is_admin)
    )

async def on_startup(dispatcher):
    await bot.delete_webhook(drop_pending_updates=True)
    from aiogram.types import BotCommand, BotCommandScopeDefault
    commands = [
        BotCommand("start",  "🏠 القائمة الرئيسية / Main Menu"),
        BotCommand("rates",  "💹 أسعار الصرف / Exchange Rates"),
        BotCommand("gold",   "🥇 أسعار الذهب / Gold Prices"),
        BotCommand("about",  "ℹ️ عن البوت / About"),
        BotCommand("help",    "❓ المساعدة / Help"),
        BotCommand("tracker", "📊 Crypto Tracker"),

        BotCommand("stats",  "📊 الإحصائيات / Statistics (Admin)"),
        BotCommand("admin",  "⚙️ لوحة التحكم / Admin Panel"),
    ]
    await bot.set_my_commands(commands, scope=BotCommandScopeDefault())
    logger.info("✅ Bot v4.1 (with Crypto Tracker) is running!")

if __name__ == "__main__":
    init_db()
    logger.info("Transfer Bot v4.1 (with Crypto Tracker) starting…")

    # ── Register Crypto Tracker handlers ─────────────────────────
    dp.register_message_handler(cmd_tracker, commands=["tracker"], state="*")
    dp.register_message_handler(
        handle_portfolio_qty,
        state=PortfolioStates.entering_quantity
    )
    dp.register_callback_query_handler(
        cb_trk_list,
        lambda c: c.data == "trk:list",
        state="*"
    )
    dp.register_callback_query_handler(
        cb_trk_token,
        lambda c: c.data.startswith("trk:token:"),
        state="*"
    )
    dp.register_callback_query_handler(
        cb_trk_ai,
        lambda c: c.data.startswith("trk:ai:"),
        state="*"
    )
    dp.register_callback_query_handler(
        cb_trk_gpt,
        lambda c: c.data.startswith("trk:gpt:"),
        state="*"
    )
    dp.register_callback_query_handler(
        cb_trk_portfolio,
        lambda c: c.data == "trk:portfolio",
        state="*"
    )
    dp.register_callback_query_handler(
        cb_pf_edit,
        lambda c: c.data.startswith("trk:pf_edit:"),
        state="*"
    )
    dp.register_callback_query_handler(
        cb_trk_admin,
        lambda c: c.data == "trk:admin",
        state="*"
    )
    dp.register_callback_query_handler(
        cb_admin_toggle,
        lambda c: c.data.startswith("trk:admin_toggle:"),
        state="*"
    )
    dp.register_callback_query_handler(
        cb_admin_add_popular,
        lambda c: c.data == "trk:admin_add_popular",
        state="*"
    )
    dp.register_callback_query_handler(
        cb_admin_add_token,
        lambda c: c.data.startswith("trk:admin_add:") and len(c.data.split(":")) >= 4,
        state="*"
    )
    dp.register_callback_query_handler(
        cb_trk_export,
        lambda c: c.data == "trk:export",
        state="*"
    )
    dp.register_callback_query_handler(
        cb_trk_close,
        lambda c: c.data == "trk:close",
        state="*"
    )
    dp.register_callback_query_handler(
        cb_admin_search_prompt,
        lambda c: c.data == "trk:admin_search",
        state="*"
    )
    dp.register_callback_query_handler(
        cb_search_already,
        lambda c: c.data.startswith("trk:search_already:"),
        state="*"
    )
    dp.register_message_handler(
        on_admin_token_search,
        state=AdminTokenSearch.waiting_query
    )
    executor.start_polling(
        dp,
        skip_updates=True,
        on_startup=on_startup,
    )
